from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from copy import copy
from pathlib import Path
import unicodedata
import re
import warnings
import sys

BASE_DIR = Path(__file__).resolve().parent

INPUT_FILE = BASE_DIR / "RESULTADOS USUARIOS 2M_Illpa_2.0.xlsx"
TARGET_FILE = BASE_DIR / "Software_Mejorado_Cultivos Anuales_2025-2026_Arapa.xlsx"
IMAGE_DIR = BASE_DIR / "extracted_images"

TARGET_SHEET = "Base_Datos"
TARGET_HEADER_ROW = 2
TARGET_ROW_TO_REPLACE = 3

# Change this to make all reinserted images smaller/larger
# 1.00 = original size
# 0.75 = 75%
# 0.60 = 60%
# 0.50 = 50%
IMAGE_SCALE = 0.40

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


ALLOWED_PRODUCTS = [
    "ACELGA",
    "AGUAYMANTO",
    "AJI",
    "AJO",
    "ALCACHOFA",
    "ALFALFA",
    "RYE GRASS",
    "ALFALFA +RYE GRASS",
    "ALGODON",
    "APIO",
    "ARROZ",
    "ARVEJA",
    "AVENA + VICIA",
    "BETARRAGA",
    "BROCOLI",
    "CACAO",
    "CAFE",
    "CAIGUA",
    "CALABAZA",
    "CAMOTE",
    "CAÑA DE AZUCAR",
    "CEBADA",
    "CEBOLLA",
    "CENTENO",
    "CIRUELO",
    "CITRICOS",
    "COCOTERO",
    "COL",
    "COLIFLOR",
    "DURAZNO",
    "ESPARRAGO",
    "ESPINACA",
    "FRESA",
    "FRIJOL",
    "GARBANZO",
    "GIRASOL",
    "GUAYABO",
    "HABA",
    "HIGUERA",
    "LECHUGA",
    "MAIZ MORADO",
    "MAIZ AMILACEO (GRANO)",
    "MANGO",
    "MANI",
    "MANZANO Y PERAL",
    "MELON",
    "NABO",
    "OLIVO",
    "PALMA ACEITERA",
    "PALTO",
    "PAPA NATIVA",
    "PAPA MEJORADA",
    "PAPAYA",
    "PASTOS ASOCIADOS",
    "PEPINO",
    "PIÑA",
    "PLATANO",
    "RABANO",
    "SANDIA",
    "SOYA",
    "TABACO",
    "TARA",
    "TOMATE",
    "TREBOL",
    "TRIGO",
    "VID",
    "ZANAHORIA",
]


if len(sys.argv) < 3:
    print('Usage: python3 fill.py "Full Name" "CROP"')
    print()
    print("Example:")
    print('  python3 fill.py "Hancco Suca Isaac" "palto"')
    sys.exit(1)

PERSON_NAME = sys.argv[1]
USER_PRODUCT_RAW = sys.argv[2].strip()


def norm(text):
    if text is None:
        return ""

    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("\n", " ")
    return text.strip()


def safe_filename(text):
    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^A-Za-z0-9_\-]", "", text)
    return text


def to_float(value):
    if value is None:
        return None

    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return None


# ------------------------------------------------------------
# 0. Check user crop first
# ------------------------------------------------------------

allowed_lookup = {norm(product): product for product in ALLOWED_PRODUCTS}
user_product_key = norm(USER_PRODUCT_RAW)

if user_product_key not in allowed_lookup:
    print(f'\nCultivo no disponible: "{USER_PRODUCT_RAW}"')
    print("\nCultivos disponibles:")
    for product in ALLOWED_PRODUCTS:
        print(f"  - {product}")
    sys.exit(1)

USER_PRODUCT = allowed_lookup[user_product_key]

OUTPUT_FILE = BASE_DIR / (
    "Software_Mejorado_Cultivos_Anuales_2025-2026_Arapa_"
    f"{safe_filename(PERSON_NAME)}_{safe_filename(USER_PRODUCT)}.xlsx"
)


def find_header_row(ws, required_header="NOMBRES Y APELLIDOS", max_rows=30):
    required = norm(required_header)

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, values_only=True),
        start=1,
    ):
        normalized_row = [norm(cell) for cell in row]

        if required in normalized_row:
            return row_index, row

    raise ValueError(f"Could not find header row with: {required_header}")


def get_headers_from_values(header_row_values):
    headers = {}

    for index, value in enumerate(header_row_values, start=1):
        if value is not None:
            key = norm(value)

            if key not in headers:
                headers[key] = []

            headers[key].append(index)

    return headers


def get_headers_from_sheet(ws, header_row):
    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value

        if value is not None:
            key = norm(value)

            if key not in headers:
                headers[key] = []

            headers[key].append(col)

    return headers


def get(row_values, headers, header_name):
    cols = headers.get(norm(header_name))

    if cols is None:
        return None

    for col in cols:
        if col - 1 < len(row_values):
            value = row_values[col - 1]

            if value not in (None, ""):
                return value

    return None


def find_person_row(ws, headers, name, start_row):
    name_cols = headers.get(norm("NOMBRES Y APELLIDOS"))

    if name_cols is None:
        raise ValueError("Column 'NOMBRES Y APELLIDOS' was not found.")

    name_col = name_cols[0]
    possible_matches = []

    for row_index, row in enumerate(
        ws.iter_rows(min_row=start_row + 1, values_only=True),
        start=start_row + 1,
    ):
        cell_value = row[name_col - 1] if name_col - 1 < len(row) else None

        if norm(cell_value) == norm(name):
            return row_index, row

        if norm(name) in norm(cell_value) or norm(cell_value) in norm(name):
            possible_matches.append((row_index, cell_value))

    print(f"\nPersona no encontrada exactamente: {name}")

    if possible_matches:
        print("\nPosibles coincidencias:")
        for row_index, value in possible_matches:
            print(f"  Row {row_index}: {value}")

    raise ValueError("Person not found.")


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if src.has_style:
            dst._style = copy(src._style)

        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def set_target_value(ws, target_headers, target_header, value, target_row):
    cols = target_headers.get(norm(target_header))

    if cols is None:
        print(f"Target header not found: {target_header}")
        return False

    col = cols[0]
    ws.cell(target_row, col).value = value
    return True


def reinsert_images(wb):
    """
    Reinserts normal PNG/JPEG images that openpyxl may remove when saving.
    The positions came from scan_image_positions.py.
    Images are resized using IMAGE_SCALE.
    """

    images_by_sheet = {
        "Textura": [
            ("image2.png", "F1"),
        ],

        "Interpretación": [
            ("image3.jpeg", "A1"),
            ("image4.png", "AF1"),
            ("image5.jpeg", "R1"),
        ],

        "Gráfico_Int": [
            ("image3.jpeg", "A1"),
            ("image4.png", "AD1"),
            ("image5.jpeg", "T1"),
        ],

        "Nec_fert": [
            ("image3.jpeg", "A1"),
            ("image5.jpeg", "E1"),
            ("image4.png", "I1"),
        ],

        "Rec_fert": [
            ("image3.jpeg", "A1"),
            ("image5.jpeg", "D1"),
            ("image4.png", "H1"),
        ],
    }

    for sheet_name, images in images_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet not found for image insertion: {sheet_name}")
            continue

        ws = wb[sheet_name]

        # Prevent duplicate images if openpyxl preserved some.
        ws._images = []

        for image_filename, anchor in images:
            image_path = IMAGE_DIR / image_filename

            if not image_path.exists():
                print(f"Image file not found: {image_path}")
                continue

            img = Image(str(image_path))

            # Resize image
            img.width = img.width * IMAGE_SCALE
            img.height = img.height * IMAGE_SCALE

            ws.add_image(img, anchor)
            print(
                f"Inserted {image_filename} into {sheet_name}!{anchor} "
                f"at {int(IMAGE_SCALE * 100)}% size"
            )


# ------------------------------------------------------------
# 1. Read input file and find person
# ------------------------------------------------------------

input_wb = load_workbook(INPUT_FILE, data_only=True, read_only=True)
input_ws = input_wb.active

input_header_row_number, input_header_row_values = find_header_row(input_ws)
input_headers = get_headers_from_values(input_header_row_values)

person_row_number, person_row_values = find_person_row(
    input_ws,
    input_headers,
    PERSON_NAME,
    input_header_row_number,
)

input_cultivo = get(person_row_values, input_headers, "CULTIVO A INSTALAR")
ph_value = get(person_row_values, input_headers, "pH")
p_value = get(person_row_values, input_headers, "P_mg/kg)")

ph_number = to_float(ph_value)
p_number = to_float(p_value)


# ------------------------------------------------------------
# 1.1 Display all relevant input values
# ------------------------------------------------------------

input_display_headers = [
    "Nº",
    "DEP",
    "PROV",
    "DIST",
    "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
    "NOMBRES Y APELLIDOS",
    "DNI",
    "TELÉFONO",
    "FECHA DE MUESTREO",
    "HORA DE MUESTREO",
    "CULTIVO ANTERIOR",
    "VARIEDAD (ANTERIOR)",
    "CULTIVO A INSTALAR",
    "VARIEDAD (A INSTALAR)",
    "HECTÁREAS",
    "COORDENADAS (E)",
    "COORDENADAS (N)",
    "ALTITUD (m.s.n.m.)",
    "CODIGO",
    "LABORATORIO",
    "ZONA",
    "ESTADO",
    "Nº DE COTIZACION",
    "INSTITUCION U ORGANIZACION",
    "RESPONSABLE",
    "CORREO/CEL",
    "OBS",
    "pH",
    "CE_mS/m",
    "CaCO3 _% Equivalente",
    "Aluminio Intercambiable cmol(+)/Kg",
    "Acidez (H+) cmol(+)/Kg",
    "MO_%",
    "Arena",
    "Arcilla",
    "Limo",
    "Clase Textural",
    "Calcio (Ca) (*)cmol(+)/Kg",
    "Magnesio (Mg) (*) cmol(+)/Kg",
    "Sodio (Na) (*) cmol(+)/Kg",
    "Potasio (K) (*) cmol(+)/Kg",
    "N_%",
    "K_ppm",
    "P_mg/kg)",
    "CICe (*) cmol(+)/Kg",
    "CICe (*) cmol(+)/Kg",
]

print(f"\nFound person: {PERSON_NAME}")
print(f"Input Excel row: {person_row_number}")
print("=" * 70)
print("INPUT FILE VALUES")
print("=" * 70)

for header in input_display_headers:
    value = get(person_row_values, input_headers, header)
    print(f"{header}: {value}")

print("=" * 70)
print(f"CULTIVO A INSTALAR from input file: {input_cultivo}")
print(f"Plan de Recomendación de Fertilización from user: {USER_PRODUCT}")
print(f"pH from input file: {ph_value}")
print(f"P_mg/kg) from input file: {p_value}")

if ph_number is None:
    print("Phosphorus method: pH invalid or empty")
elif ph_number < 7.0:
    print("Phosphorus method: Bray because pH < 7.0")
else:
    print("Phosphorus method: Olsen because pH >= 7.0")

print("=" * 70)


# ------------------------------------------------------------
# 2. Open target file using openpyxl
# ------------------------------------------------------------

target_wb = load_workbook(TARGET_FILE)

if TARGET_SHEET not in target_wb.sheetnames:
    print("Available sheets:")
    for sheet in target_wb.sheetnames:
        print(repr(sheet))
    raise KeyError(f"Target sheet not found: {TARGET_SHEET}")

target_ws = target_wb[TARGET_SHEET]
target_headers = get_headers_from_sheet(target_ws, TARGET_HEADER_ROW)

target_row_number = TARGET_ROW_TO_REPLACE

copy_row_style(target_ws, TARGET_ROW_TO_REPLACE, target_row_number)

print(f"Writing to target sheet: {TARGET_SHEET}")
print(f"Replacing target Excel row: {target_row_number}")
print("=" * 70)


# ------------------------------------------------------------
# 3. Mapping: target column -> input column
# ------------------------------------------------------------

mapping = {
    "CÓDIGO": "CODIGO",
    "Código Laboratorio": "CODIGO",
    "Cotizacion Servicio": "Nº DE COTIZACION",
    "Fecha Muestreo": "FECHA DE MUESTREO",
    "Hora Muestreo": "HORA DE MUESTREO",

    "Codigo_Muestra_Cliente": "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
    "Cliente": "INSTITUCION U ORGANIZACION",
    "Propietario / Productor": "NOMBRES Y APELLIDOS",
    "Direccion del Cliente": "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
    "Solicitado por": "RESPONSABLE",
    "Muestreado por": "NOMBRES Y APELLIDOS",

    "Procedencia de la Muestra": "DIST",

    "Acidez Intercambiable": "Acidez (H+) cmol(+)/Kg",
    "Aluminio Intercambiable": "Aluminio Intercambiable cmol(+)/Kg",

    "Ca_Cmol/Kg": "Calcio (Ca) (*)cmol(+)/Kg",
    "Mg_Cmol/Kg": "Magnesio (Mg) (*) cmol(+)/Kg",
    "Na_Cmol/Kg": "Sodio (Na) (*) cmol(+)/Kg",
    "K_Cmol/Kg": "Potasio (K) (*) cmol(+)/Kg",

    "Calcio Intercambiable": "Calcio (Ca) (*)cmol(+)/Kg",
    "Magnesio Intercambiable": "Magnesio (Mg) (*) cmol(+)/Kg",
    "Sodio Intercambiable": "Sodio (Na) (*) cmol(+)/Kg",
    "Potasio Intercambiable": "Potasio (K) (*) cmol(+)/Kg",

    "Carbonato de Calcio Equivalente": "CaCO3 _% Equivalente",
    "Materia Orgánica (AS-07 Método de Walkley y Black)": "MO_%",
    "Materia Organica (AS-07 Walkley y Black)": "MO_%",
    "Materia Orgánica por LECO": "MO_%",

    "Conductividad Electrica (Suelo)": "CE_mS/m",
    "pH. (Suelo)": "pH",

    "Fósforo Disponible (Bray y Kurtz)": "P_mg/kg)",
    "Fósforo Disponible Bray mpaes": "P_mg/kg)",
    "Potasio Disponible (MPAES)": "K_ppm",
    "Potasio Disponible (AA)": "K_ppm",

    "Arena": "Arena",
    "Arcilla": "Arcilla",
    "Limo": "Limo",
    "Clase Textural": "Clase Textural",

    "Nitrógeno Total": "N_%",
    "Nitrógeno Total Kjeldahl": "N_%",
}


# ------------------------------------------------------------
# 4. Write mapped values from input file
#    Missing or empty mapped values are written as 0.
# ------------------------------------------------------------

for target_header, input_header in mapping.items():
    value = get(person_row_values, input_headers, input_header)

    if value is None:
        value = 0
        print(f"{target_header}: INPUT HEADER NOT FOUND or EMPTY -> set to 0")

    set_target_value(
        target_ws,
        target_headers,
        target_header,
        value,
        target_row_number,
    )

    print(f"{target_header}: {value}")


# ------------------------------------------------------------
# 5. Fixed values and user-selected crop
# ------------------------------------------------------------

fixed_values = {
    "Número de Muestras": 1,
    "Presentación Muestra": "1 kg",
    "Plan de Recomendación de Fertilización": USER_PRODUCT,
}

for target_header, value in fixed_values.items():
    ok = set_target_value(
        target_ws,
        target_headers,
        target_header,
        value,
        target_row_number,
    )

    if ok:
        print(f"{target_header}: {value}")


# ------------------------------------------------------------
# 6. Force specific target cells
# ------------------------------------------------------------

# T3 is Plan de Recomendación de Fertilización.
target_ws["T3"] = USER_PRODUCT
print(f"T3 / Plan de Recomendación de Fertilización: {USER_PRODUCT}")

# Phosphorus method based on pH:
# AL3 = Fósforo Disponible (Olsen)
# AM3 = Fósforo Disponible (Bray y Kurtz)
#
# if pH < 7.0  -> Bray  -> AM3
# if pH >= 7.0 -> Olsen -> AL3
target_ws["AL3"] = 0
target_ws["AM3"] = 0

if p_number is None:
    print("P_mg/kg) not found or invalid. AL3 and AM3 set to 0.")

elif ph_number is None:
    print("pH not found or invalid. Cannot choose Olsen/Bray. AL3 and AM3 set to 0.")

elif ph_number < 7.0:
    target_ws["AM3"] = p_number
    print(f"pH: {ph_number}")
    print("Using Bray because pH < 7.0")
    print("AL3 / Fósforo Disponible (Olsen): 0")
    print(f"AM3 / Fósforo Disponible (Bray y Kurtz): {p_number}")

else:
    target_ws["AL3"] = p_number
    print(f"pH: {ph_number}")
    print("Using Olsen because pH >= 7.0")
    print(f"AL3 / Fósforo Disponible (Olsen): {p_number}")
    print("AM3 / Fósforo Disponible (Bray y Kurtz): 0")


# ------------------------------------------------------------
# 7. Reinsert images into sheets where openpyxl may remove them
# ------------------------------------------------------------

reinsert_images(target_wb)


# ------------------------------------------------------------
# 8. Save output
# ------------------------------------------------------------

target_wb.save(OUTPUT_FILE)

print("=" * 70)
print("Done.")
print(f"Saved as: {OUTPUT_FILE.name}")
