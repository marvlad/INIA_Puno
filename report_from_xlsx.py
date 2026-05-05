from openpyxl import load_workbook
from pathlib import Path
import unicodedata
import re
import sys
import warnings
import argparse
import shutil
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    HRFlowable, Image as RLImage,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "RESULTADOS USUARIOS 2M_Illpa_2.0.xlsx"
IMAGE_DIR = BASE_DIR / "extracted_images"

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Generate PDF soil report from Excel data."
    )
    parser.add_argument("--name", help="Full name to search in the Excel file")
    parser.add_argument("--cultivo", help="Tipo de cultivo / crop")
    parser.add_argument("--output", help="Output PDF path")
    parser.add_argument("old_name", nargs="?", help="Old style full name")
    parser.add_argument("old_cultivo", nargs="?", help="Old style crop")

    args = parser.parse_args()
    name = args.name or args.old_name
    cultivo = args.cultivo or args.old_cultivo

    if not name or not cultivo:
        print("Usage:")
        print('  python3 report_pdf.py --name "Full Name" --cultivo "CROP" --output output/report.pdf')
        print('  python3 report_pdf.py "Full Name" "CROP"')
        sys.exit(1)

    return args, name.strip(), cultivo.strip()


ARGS, PERSON_NAME, USER_PRODUCT_RAW = parse_arguments()

# ---------------------------------------------------------------------------
# Allowed crops
# ---------------------------------------------------------------------------

ALLOWED_PRODUCTS = [
    "ACELGA", "AGUAYMANTO", "AJI", "AJO", "ALCACHOFA", "ALFALFA",
    "RYE GRASS", "ALFALFA +RYE GRASS", "ALGODON", "APIO", "ARROZ",
    "ARVEJA", "AVENA + VICIA", "BETARRAGA", "BROCOLI", "CACAO",
    "CAFE", "CAIGUA", "CALABAZA", "CAMOTE", "CAÑA DE AZUCAR",
    "CEBADA", "CEBOLLA", "CENTENO", "CIRUELO", "CITRICOS",
    "COCOTERO", "COL", "COLIFLOR", "DURAZNO", "ESPARRAGO",
    "ESPINACA", "FRESA", "FRIJOL", "GARBANZO", "GIRASOL",
    "GUAYABO", "HABA", "HIGUERA", "LECHUGA", "MAIZ MORADO",
    "MAIZ AMILACEO (GRANO)", "MANGO", "MANI", "MANZANO Y PERAL",
    "MELON", "NABO", "OLIVO", "PALMA ACEITERA", "PALTO",
    "PAPA NATIVA", "PAPA MEJORADA", "PAPAYA", "PASTOS ASOCIADOS",
    "PEPINO", "PIÑA", "PLATANO", "RABANO", "SANDIA", "SOYA",
    "TABACO", "TARA", "TOMATE", "TREBOL", "TRIGO", "VID", "ZANAHORIA",
]

# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def norm(text):
    if text is None:
        return ""
    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("\n", " ")
    return text.strip()


def safe_filename(text):
    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^A-Za-z0-9_\-]", "", text)
    return text or "reporte"


def to_float(value):
    if value is None:
        return None
    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return None


def fmt(value, decimals=2):
    x = to_float(value)
    if x is None:
        return "0" if value is None else str(value)
    return f"{x:.{decimals}f}"


def fmt_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value)


def s(v):
    """Safe string – return empty string if None."""
    return "" if v is None else str(v)


# ---------------------------------------------------------------------------
# Validate crop
# ---------------------------------------------------------------------------

allowed_lookup = {norm(p): p for p in ALLOWED_PRODUCTS}
user_product_key = norm(USER_PRODUCT_RAW)

if user_product_key not in allowed_lookup:
    print(f'\nCultivo no disponible: "{USER_PRODUCT_RAW}"')
    print("\nCultivos disponibles:")
    for p in ALLOWED_PRODUCTS:
        print(f"  - {p}")
    sys.exit(1)

USER_PRODUCT = allowed_lookup[user_product_key]

BASE_OUTPUT = f"Informe_{safe_filename(PERSON_NAME)}_{safe_filename(USER_PRODUCT)}"

if ARGS.output:
    OUTPUT_PDF = Path(ARGS.output).resolve()
    OUTPUT_PDF.parent.mkdir(parents=True, exist_ok=True)
else:
    OUTPUT_PDF = BASE_DIR / f"{BASE_OUTPUT}.pdf"

# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------

def find_header_row(ws, required_header="NOMBRES Y APELLIDOS", max_rows=30):
    required = norm(required_header)
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1
    ):
        if required in [norm(cell) for cell in row]:
            return row_index, row
    raise ValueError(f"Could not find header row with: {required_header}")


def get_headers_from_values(header_row_values):
    headers = {}
    for index, value in enumerate(header_row_values, start=1):
        if value is not None:
            key = norm(value)
            headers.setdefault(key, []).append(index)
    return headers


def get(row_values, headers, header_name):
    cols = headers.get(norm(header_name))
    if cols is None:
        return None
    for col in cols:
        if col - 1 < len(row_values):
            value = row_values[col - 1]
            if value not in (None, ""):
                return value
    return None


def find_person_row(ws, headers, name, start_row):
    name_cols = headers.get(norm("NOMBRES Y APELLIDOS"))
    if name_cols is None:
        raise ValueError("Column 'NOMBRES Y APELLIDOS' was not found.")
    name_col = name_cols[0]
    possible_matches = []

    for row_index, row in enumerate(
        ws.iter_rows(min_row=start_row + 1, values_only=True), start=start_row + 1
    ):
        cell_value = row[name_col - 1] if name_col - 1 < len(row) else None
        if norm(cell_value) == norm(name):
            return row_index, row
        if norm(name) in norm(cell_value) or norm(cell_value) in norm(name):
            possible_matches.append((row_index, cell_value))

    print(f"\nPersona no encontrada exactamente: {name}")
    if possible_matches:
        print("\nPosibles coincidencias:")
        for row_index, value in possible_matches:
            print(f"  Row {row_index}: {value}")
    raise ValueError("Person not found.")


def read_person_data():
    wb = load_workbook(INPUT_FILE, data_only=True, read_only=True)
    ws = wb.active
    header_row_number, header_row_values = find_header_row(ws)
    headers = get_headers_from_values(header_row_values)
    person_row_number, person_row_values = find_person_row(
        ws, headers, PERSON_NAME, header_row_number
    )
    data = {
        "row": person_row_number,
        "name": get(person_row_values, headers, "NOMBRES Y APELLIDOS"),
        "dep": get(person_row_values, headers, "DEP"),
        "prov": get(person_row_values, headers, "PROV"),
        "dist": get(person_row_values, headers, "DIST"),
        "localidad": get(person_row_values, headers, "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)"),
        "fecha_muestreo": get(person_row_values, headers, "FECHA DE MUESTREO"),
        "cultivo_input": get(person_row_values, headers, "CULTIVO A INSTALAR"),
        "variedad_instalar": get(person_row_values, headers, "VARIEDAD (A INSTALAR)"),
        "hectareas": get(person_row_values, headers, "HECTÁREAS"),
        "altitud": get(person_row_values, headers, "ALTITUD (m.s.n.m.)"),
        "codigo": get(person_row_values, headers, "CODIGO"),
        "cotizacion": get(person_row_values, headers, "Nº DE COTIZACION"),
        "institucion": get(person_row_values, headers, "INSTITUCION U ORGANIZACION"),
        "responsable": get(person_row_values, headers, "RESPONSABLE"),
        "ph": get(person_row_values, headers, "pH"),
        "ce": get(person_row_values, headers, "CE_mS/m"),
        "caco3": get(person_row_values, headers, "CaCO3 _% Equivalente"),
        "al": get(person_row_values, headers, "Aluminio Intercambiable cmol(+)/Kg"),
        "acidez": get(person_row_values, headers, "Acidez (H+) cmol(+)/Kg"),
        "mo": get(person_row_values, headers, "MO_%"),
        "arena": get(person_row_values, headers, "Arena"),
        "arcilla": get(person_row_values, headers, "Arcilla"),
        "limo": get(person_row_values, headers, "Limo"),
        "clase_textural": get(person_row_values, headers, "Clase Textural"),
        "ca": get(person_row_values, headers, "Calcio (Ca) (*)cmol(+)/Kg"),
        "mg": get(person_row_values, headers, "Magnesio (Mg) (*) cmol(+)/Kg"),
        "na": get(person_row_values, headers, "Sodio (Na) (*) cmol(+)/Kg"),
        "k": get(person_row_values, headers, "Potasio (K) (*) cmol(+)/Kg"),
        "n": get(person_row_values, headers, "N_%"),
        "k_ppm": get(person_row_values, headers, "K_ppm"),
        "p": get(person_row_values, headers, "P_mg/kg)"),
        "cice": get(person_row_values, headers, "CICe (*) cmol(+)/Kg"),
    }
    return data

# ---------------------------------------------------------------------------
# Interpretation helpers
# ---------------------------------------------------------------------------

def interp_ph(value):
    x = to_float(value)
    if x is None: return "Sin dato"
    if x < 5.5:  return "Fuertemente acido"
    if x < 6.5:  return "Moderadamente acido"
    if x <= 7.3: return "Neutro"
    if x <= 8.0: return "Moderadamente alcalino"
    return "Fuertemente alcalino"

def interp_ce(value):
    x = to_float(value)
    if x is None:  return "Sin dato"
    if x < 100:    return "Normal"
    if x < 200:    return "Muy ligeramente salino"
    if x < 400:    return "Moderadamente salino"
    if x < 800:    return "Suelo salino"
    return "Fuertemente salino"

def interp_caco3(value):
    x = to_float(value)
    if x is None: return "Sin dato"
    if x < 1:     return "Sin problema - Bajo contenido de carbonatos"
    if x < 5:     return "Bajo contenido de carbonatos"
    if x < 15:    return "Contenido medio de carbonatos"
    return "Alto contenido de carbonatos"

def interp_mo(value):
    x = to_float(value)
    if x is None: return "Sin dato"
    if x < 2:     return "Baja disponibilidad"
    if x < 4:     return "Media disponibilidad"
    return "Alta disponibilidad"

def interp_n(value):
    x = to_float(value)
    if x is None:   return "Sin dato"
    if x < 0.10:    return "Bajo"
    if x <= 0.15:   return "Medio"
    return "Alto"

def interp_p(value):
    x = to_float(value)
    if x is None: return "Sin dato"
    if x < 5.5:   return "Muy bajo"
    if x <= 11:   return "Bajo"
    if x <= 20:   return "Medio"
    if x <= 40:   return "Alto"
    return "Muy alto"

def interp_k(value):
    x = to_float(value)
    if x is None: return "Sin dato"
    if x < 120:   return "Bajo"
    if x <= 240:  return "Medio"
    return "Alto"

def saturation_values(data):
    ca = to_float(data["ca"]) or 0
    mg = to_float(data["mg"]) or 0
    k  = to_float(data["k"])  or 0
    na = to_float(data["na"]) or 0
    al = to_float(data["al"]) or 0
    total = ca + mg + k + na + al
    if total == 0:
        return {"Calcio": 0, "Magnesio": 0, "Potasio": 0, "Sodio": 0, "Aluminio": 0}
    return {
        "Calcio":   ca / total * 100,
        "Magnesio": mg / total * 100,
        "Potasio":  k  / total * 100,
        "Sodio":    na / total * 100,
        "Aluminio": al / total * 100,
    }

def relations(data):
    ca = to_float(data["ca"]) or 0
    mg = to_float(data["mg"]) or 0
    k  = to_float(data["k"])  or 0
    return {
        "ca_k":    ca / k if k else 0,
        "mg_k":    mg / k if k else 0,
        "ca_mg_k": (ca + mg) / k if k else 0,
        "ca_mg":   ca / mg if mg else 0,
    }

# ---------------------------------------------------------------------------
# Drawing helpers
# ---------------------------------------------------------------------------

# Color map for scale bars
COLOR_CYAN   = colors.Color(0, 0.71, 0.86)
COLOR_YELLOW = colors.yellow
COLOR_GREEN  = colors.green
COLOR_BLUE   = colors.Color(0, 0.2, 0.7)
COLOR_RED    = colors.red
COLOR_GRAY   = colors.grey
COLOR_HEADER = colors.Color(0.90, 0.90, 0.86)
COLOR_LABGRAY= colors.Color(0.82, 0.82, 0.82)
COLOR_SKY    = colors.Color(0, 0.71, 0.86)
COLOR_DEEPBLUE = colors.Color(0, 0.20, 0.71)

def bar_svg(value, max_value, bar_color, bar_width_pt=200, bar_height_pt=8):
    """Return a small Table that visually mimics a horizontal bar."""
    x = to_float(value)
    pct = max(0, min(1.0, (x or 0) / max_value)) if max_value else 0
    filled = bar_width_pt * pct
    empty  = bar_width_pt - filled
    data = [[""]]
    style = [
        ("BACKGROUND", (0, 0), (0, 0), bar_color),
        ("GRID",       (0, 0), (-1, -1), 0, colors.transparent),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]
    bar_table = Table(data, colWidths=[filled or 0.1], rowHeights=[bar_height_pt])
    bar_table.setStyle(TableStyle(style))
    # Wrap in outer table with remaining space
    outer = Table([[bar_table, ""]], colWidths=[filled or 0.1, empty or 0.1], rowHeights=[bar_height_pt])
    outer.setStyle(TableStyle([
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("GRID",          (0, 0), (-1, -1), 0, colors.transparent),
    ]))
    return outer


def color_scale_table(labels_colors, col_width=None):
    """Horizontal color-coded scale row with tiny labels."""
    n = len(labels_colors)
    cw = col_width or (200 / n)
    row_labels = [lbl for lbl, _ in labels_colors]
    row_colors = [c for _, c in labels_colors]

    label_row = [Paragraph(f'<font size="5">{lbl}</font>', _center_style()) for lbl in row_labels]
    t = Table([label_row], colWidths=[cw] * n, rowHeights=[12])
    style_cmds = [
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.white),
        ("TOPPADDING",    (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("LEFTPADDING",   (0, 0), (-1, -1), 1),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 1),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]
    for i, c in enumerate(row_colors):
        style_cmds.append(("BACKGROUND", (i, 0), (i, 0), c))
    t.setStyle(TableStyle(style_cmds))
    return t


def _bold_style(size=7, align=TA_LEFT):
    return ParagraphStyle("bold", fontName="Helvetica-Bold", fontSize=size, alignment=align, leading=size + 2)

def _normal_style(size=7, align=TA_LEFT):
    return ParagraphStyle("normal", fontName="Helvetica", fontSize=size, alignment=align, leading=size + 2)

def _center_style(size=6):
    return ParagraphStyle("center", fontName="Helvetica", fontSize=size, alignment=TA_CENTER, leading=size + 2)

def _header_style(size=9):
    return ParagraphStyle("header", fontName="Helvetica-Bold", fontSize=size, alignment=TA_CENTER, leading=size + 3)

# ---------------------------------------------------------------------------
# PDF builder
# ---------------------------------------------------------------------------

def build_pdf(data):
    doc = SimpleDocTemplate(
        str(OUTPUT_PDF),
        pagesize=A4,
        rightMargin=0.8 * cm,
        leftMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )

    story = []

    # ---- HEADER logos ----
    logo_left   = IMAGE_DIR / "image3.jpeg"
    logo_center = IMAGE_DIR / "image4.png"
    logo_right  = IMAGE_DIR / "image5.jpeg"

    LOGO_HEIGHT = 1.0 * cm

    def logo_cell(path, height=LOGO_HEIGHT):
        if Path(path).exists():
            # Load image to get natural dimensions and compute width that preserves aspect ratio
            from PIL import Image as PILImage
            with PILImage.open(str(path)) as pil_img:
                nat_w, nat_h = pil_img.size
            aspect = nat_w / nat_h
            img = RLImage(str(path), width=height * aspect, height=height)
            img.hAlign = "CENTER"
            return img
        return Paragraph(path.name, _normal_style())

    logo_l = logo_cell(logo_left)
    logo_c = logo_cell(logo_center)
    logo_r = logo_cell(logo_right)

    header_table = Table(
        [[logo_l, logo_c, logo_r]],
        colWidths=["33%", "34%", "33%"],
    )
    header_table.setStyle(TableStyle([
        ("ALIGN",  (0, 0), (0, 0), "LEFT"),
        ("ALIGN",  (1, 0), (1, 0), "CENTER"),
        ("ALIGN",  (2, 0), (2, 0), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",   (0, 0), (-1, -1), 0, colors.transparent),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(HRFlowable(width="100%", thickness=3, color=colors.black))
    story.append(Spacer(1, 6))

    # ---- INFORMACION GENERAL title ----
    story.append(Paragraph("INFORMACION GENERAL", _header_style(size=11)))
    story.append(Spacer(1, 4))

    bold7  = _bold_style(7)
    norm7  = _normal_style(7)

    def field(label, value):
        return [Paragraph(label, bold7), Paragraph(s(value), norm7)]

    general_data = [
        field("NOMBRE AGRICULTOR", data["name"]) +
        field("FECHA RECEPCION", fmt_date(data["fecha_muestreo"])),

        field("DIRECCION", data["dist"]) +
        field("FECHA DE ANALISIS", ""),

        field("NOMBRE PARCELA", data["localidad"]) +
        field("FECHA DE EMISION", ""),

        field("ALTITUD (MSNM)", data["altitud"]) +
        field("CODIGO DE LABORATORIO", data["codigo"]),

        field("AREA (HA)", data["hectareas"]) + ["", ""],

        field("CULTIVO / VARIEDAD", USER_PRODUCT) + ["", ""],
    ]

    col_w = [3.0 * cm, 6.0 * cm, 3.4 * cm, 6.0 * cm]
    gen_table = Table(general_data, colWidths=col_w)
    gen_table.setStyle(TableStyle([
        ("GRID",          (0, 0), (-1, -1), 0, colors.transparent),
        ("LINEBELOW",     (1, 0), (1, -1), 0.5, colors.black),
        ("LINEBELOW",     (3, 0), (3, -1), 0.5, colors.black),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(gen_table)

    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=3, color=colors.black))
    story.append(Spacer(1, 4))

    # ---- INTERPRETACION - ANALISIS DE SUELO ----
    story.append(Paragraph("INTERPRETACION - ANALISIS DE SUELO", _header_style(size=11)))
    story.append(Spacer(1, 4))

    page_w = A4[0] - 1.6 * cm   # usable width
    c1 = 3.2 * cm
    c2 = 1.6 * cm
    c3 = 1.8 * cm
    c4 = 2.2 * cm
    c5 = page_w - c1 - c2 - c3 - c4

    # Color scale rows for pH and CE
    ph_scale = color_scale_table([
        ("Fuertemente acido",     COLOR_RED),
        ("Moderadamente acido",   COLOR_YELLOW),
        ("Neutro",                COLOR_GREEN),
        ("Moderadamente alcalino",COLOR_CYAN),
        ("Fuertemente alcalino",  COLOR_BLUE),
    ], col_width=c5 / 5)

    ce_scale = color_scale_table([
        ("Normal",                colors.Color(0, 0.7, 0.3)),
        ("Muy ligeramente salino",COLOR_CYAN),
        ("Moderadamente salino",  colors.Color(0, 0.4, 0.7)),
        ("Suelo salino",          COLOR_BLUE),
        ("Fuertemente salino",    COLOR_DEEPBLUE),
    ], col_width=c5 / 5)

    nutrient_scale = color_scale_table([
        ("Muy bajo", COLOR_RED),
        ("Bajo",     COLOR_YELLOW),
        ("Medio",    COLOR_GREEN),
        ("Alto",     COLOR_CYAN),
        ("Muy alto", COLOR_BLUE),
    ], col_width=c5 / 5)

    B = _bold_style(7, TA_CENTER)
    N = _normal_style(7, TA_CENTER)

    def hdr(txt): return Paragraph(txt, B)
    def cel(txt): return Paragraph(str(txt), N)

    analysis_rows = [
        # sub-header
        [hdr("CODIGO DE LABORATORIO"), Paragraph(s(data["codigo"]), B), "", "", ""],
        # column headers
        [hdr("Determinacion"), hdr("Result."), hdr("Unid."), hdr("Rango adecuado"), hdr("INTERPRETACION")],
        # pH
        [cel("pH"), cel(fmt(data["ph"])), cel("Unid pH"), cel("6.6 - 7.3"), ph_scale],
        ["", "", "", "", bar_svg(data["ph"], 9, COLOR_CYAN, bar_width_pt=c5)],
        # CE
        [cel("CE"), cel(fmt(data["ce"])), cel("mS/m"), cel("<100 - 200"), ce_scale],
        ["", "", "", "", bar_svg(data["ce"], 500, COLOR_CYAN, bar_width_pt=c5)],
        # nutrient header
        [hdr("Determinacion"), hdr("Result."), hdr("Unid."), hdr("Rango adecuado"), nutrient_scale],
    ]

    nutrient_rows_def = [
        ("Materia Organica",              data["mo"],    "cmol(+)/kg", "1.6 - 3.5",    2.0,   data["mo"],   6,   COLOR_YELLOW),
        ("Nitrogeno Total",               data["mo"]*0.045,     "%",          "0.10 - 0.15",   0.15,  data["mo"]*0.045,    0.35,COLOR_YELLOW),
        ("Fosforo disponible",            data["p"],     "mg/kg",      "5.5 - 11",      11,    data["p"],    60,  COLOR_BLUE),
        ("Potasio disponible",            data["k_ppm"], "mg/kg",      "120 - 240",     240,   data["k_ppm"],500, COLOR_CYAN),
        ("Calcio intercambiable",         data["ca"],    "cmol(+)/kg", "5.0 - 10",      10,    data["ca"],   60,  COLOR_GREEN),
        ("Magnesio intercambiable",       data["mg"],    "cmol(+)/kg", "1.3 - 3.0",     3.0,   data["mg"],   10,  COLOR_YELLOW),
        ("Potasio intercambiable",        data["k"],     "cmol(+)/kg", "0.3 - 0.6",     0.6,   data["k"],    2,   COLOR_BLUE),
        ("Sodio intercambiable",          data["na"],    "cmol(+)/kg", "<0.1 - 0.3",    0.3,   data["na"],   8,   COLOR_BLUE),
        ("Aluminio intercambiable",       data["al"],    "cmol(+)/kg", "<0.1 - 0.25",   0.25,  data["al"],   1,   COLOR_GRAY),
        ("CICe",                          data["cice"],  "cmol(+)/kg", "15 - 25",        25,    data["cice"], 70,  COLOR_YELLOW),
        ("Carbonato de Calcio Equiv.",    data["caco3"], "%",          "<1 - 5",         5,     data["caco3"],15,  COLOR_YELLOW),
    ]

    for label, val, unit, rango, _, bar_val, bar_max, bar_color in nutrient_rows_def:
        analysis_rows.append([cel(label), cel(fmt(val)), cel(unit), cel(rango), ""])
        analysis_rows.append(["", "", "", "", bar_svg(bar_val, bar_max, bar_color, bar_width_pt=c5)])

    analysis_table = Table(analysis_rows, colWidths=[c1, c2, c3, c4, c5])

    row_count = len(analysis_rows)
    a_style = [
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND",    (0, 0), (-1, 0), COLOR_HEADER),
        ("SPAN",          (0, 0), (1, 0)),
        ("SPAN",          (2, 0), (4, 0)),
        ("BACKGROUND",    (0, 1), (-1, 1), COLOR_HEADER),
        ("BACKGROUND",    (0, 2), (-1, 2), COLOR_HEADER),
        ("BACKGROUND",    (0, 4), (-1, 4), COLOR_HEADER),
        ("BACKGROUND",    (0, 6), (-1, 6), COLOR_HEADER),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
    ]
    # Grey background for every label row (even rows from index 7 onward)
    for i in range(7, row_count, 2):
        a_style.append(("BACKGROUND", (0, i), (-1, i), COLOR_HEADER))

    analysis_table.setStyle(TableStyle(a_style))
    story.append(analysis_table)
    story.append(Spacer(1, 6))

    # ---- Cation relations ----
    rel = relations(data)
    rel_data = [
        [hdr("Ca/K"), hdr("Mg/K"), hdr("(Ca+Mg)/K"), hdr("Ca/Mg")],
        [cel(f'{rel["ca_k"]:.2f}'), cel(f'{rel["mg_k"]:.2f}'), cel(f'{rel["ca_mg_k"]:.2f}'), cel(f'{rel["ca_mg"]:.2f}')],
        [cel("Def. de K"), cel("Aceptable"), cel("Def. de K"), cel("Ideal")],
    ]
    rel_table = Table(rel_data, colWidths=[page_w / 4] * 4)
    rel_table.setStyle(TableStyle([
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND",    (0, 0), (-1, 0), COLOR_LABGRAY),
        ("SPAN",          (0, 0), (-1, 0)),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(rel_table)
    story.append(Spacer(1, 6))

    # ---- Saturation ----
    sat = saturation_values(data)
    sat_names  = ["Calcio", "Magnesio", "Potasio", "Sodio", "Aluminio"]
    sat_colors = [COLOR_CYAN, COLOR_RED, COLOR_CYAN, COLOR_GREEN, COLOR_GRAY]

    sat_bars = [bar_svg(sat[n], 100, sat_colors[i], bar_width_pt=2.0*cm, bar_height_pt=40)
                for i, n in enumerate(sat_names)]

    sat_data = [
        [Paragraph("Interpretacion", _bold_style(7, TA_CENTER))] + sat_bars,
        [Paragraph("% Saturacion", _bold_style(7, TA_CENTER))] +
            [cel(f'{sat[n]:.2f}') for n in sat_names],
        [Paragraph("Cationes Intercambiables", _bold_style(6, TA_CENTER))] +
            [cel(n) for n in sat_names],
    ]
    cw_sat = [3.2 * cm] + [2.5 * cm] * 5
    sat_table = Table(sat_data, colWidths=cw_sat)
    sat_table.setStyle(TableStyle([
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND",    (0, 2), (-1, 2), COLOR_LABGRAY),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(sat_table)

    # ---- Page 2: Summary ----
    story.append(PageBreak())
    story.append(Paragraph("RESUMEN DE INTERPRETACION", _header_style(size=13)))
    story.append(Spacer(1, 20))

    summary_rows = [
        [hdr("Resultados de laboratorio"), hdr(""), hdr("Interpretacion")],
        [cel("pH"),            cel(fmt(data["ph"])),              cel(interp_ph(data["ph"]))],
        [cel("C.E."),          cel(fmt(data["ce"])),              cel(interp_ce(data["ce"]))],
        [cel("CaCO3"),         cel(fmt(data["caco3"])),           cel(interp_caco3(data["caco3"]))],
        [cel("M.O."),          cel(fmt(data["mo"])),              cel(interp_mo(data["mo"]))],
        [cel("N total"),       cel(fmt(data["n"])),               cel(interp_n(data["n"]))],
        [cel("P disponible"),  cel(fmt(data["p"])),               cel(interp_p(data["p"]))],
        [cel("K disponible"),  cel(fmt(data["k_ppm"])),           cel(interp_k(data["k_ppm"]))],
        [cel("Clase textural"),cel(s(data["clase_textural"])),    cel("Textura del suelo")],
        [cel("Cultivo"),       cel(USER_PRODUCT),                 cel("Plan de recomendacion seleccionado por el usuario")],
    ]
    cw_sum = [4.0 * cm, 2.5 * cm, page_w - 6.5 * cm]
    sum_table = Table(summary_rows, colWidths=cw_sum)
    sum_style = [
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND",    (0, 0), (-1, 0), COLOR_LABGRAY),
        ("SPAN",          (0, 0), (1, 0)),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
    ]
    sum_table.setStyle(TableStyle(sum_style))
    story.append(sum_table)

    doc.build(story)
    print(f"Saved PDF: {OUTPUT_PDF}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    data = read_person_data()
    print(f"\nFound person: {data['name']}")
    print(f"Input Excel row: {data['row']}")
    print(f"CULTIVO A INSTALAR from input file: {data['cultivo_input']}")
    print(f"Plan de Recomendacion from user: {USER_PRODUCT}")
    print(f"pH: {data['ph']}")
    print(f"P_mg/kg): {data['p']}")
    build_pdf(data)


if __name__ == "__main__":
    main()
