# excel_writer.py

from pathlib import Path
import csv
import time
import subprocess

from openpyxl import load_workbook


# ------------------------------------------------------------
# Read vector from CSV
# ------------------------------------------------------------

def read_vector_from_csv(csv_file):
    """
    Reads a CSV containing fertilizer optimal values.

    It supports either:

    1) A simple one-row vector:
        10,20,30,40,50

    2) A CSV with one value per row:
        value
        10
        20
        30
        40
        50

    3) A CSV with a column named value, dose, optimal, or kg_ha.
    """
    csv_file = Path(csv_file)

    if not csv_file.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_file}")

    values = []

    with open(csv_file, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(2048)
        f.seek(0)

        # Try DictReader first
        reader = csv.DictReader(f)

        if reader.fieldnames:
            normalized_headers = [
                str(h).strip().lower()
                for h in reader.fieldnames
                if h is not None
            ]

            possible_columns = [
                "value",
                "values",
                "dose",
                "dosis",
                "optimal",
                "kg_ha",
                "kg/ha",
            ]

            selected_column = None

            for col in possible_columns:
                if col in normalized_headers:
                    selected_column = reader.fieldnames[normalized_headers.index(col)]
                    break

            if selected_column is not None:
                for row in reader:
                    raw_value = row.get(selected_column, "")

                    if raw_value is None or str(raw_value).strip() == "":
                        continue

                    values.append(float(str(raw_value).replace(",", ".").strip()))

                return values

        # If DictReader did not work, parse as normal CSV
        f.seek(0)
        reader2 = csv.reader(f)

        for row in reader2:
            for item in row:
                item = str(item).strip()

                if not item:
                    continue

                # Skip obvious text headers
                try:
                    values.append(float(item.replace(",", ".")))
                except ValueError:
                    continue

    if not values:
        raise ValueError(f"No numeric values found in CSV: {csv_file}")

    return values


# ------------------------------------------------------------
# Write optimal vector to Excel
# ------------------------------------------------------------

def write_vector_to_excel(
    excel_file,
    csv_file,
    output_excel,
    sheet_name="Nec_fert",
    start_row=53,
    column="C",
):
    """
    Writes optimal fertilizer values from CSV into Excel.

    Default target:
        Nec_fert!C53:C57

    Parameters
    ----------
    excel_file:
        Input Excel file.

    csv_file:
        CSV file containing optimal values.

    output_excel:
        Output Excel file to save.

    sheet_name:
        Sheet where values should be written.

    start_row:
        First row where values should be written.

    column:
        Excel column where values should be written.
    """
    excel_file = Path(excel_file)
    csv_file = Path(csv_file)
    output_excel = Path(output_excel)

    if not excel_file.exists():
        raise FileNotFoundError(f"Input Excel file not found: {excel_file}")

    if not csv_file.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_file}")

    values = read_vector_from_csv(csv_file)

    output_excel.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(excel_file)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet not found: {sheet_name}\n"
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    for i, value in enumerate(values):
        ws[f"{column}{start_row + i}"] = value

    wb.save(output_excel)

    print("Values written to Excel:")
    print(f"  Input Excel: {excel_file}")
    print(f"  CSV values: {csv_file}")
    print(f"  Output Excel: {output_excel}")
    print(f"  Sheet: {sheet_name}")
    print(f"  Start cell: {column}{start_row}")
    print(f"  Number of values: {len(values)}")


# ------------------------------------------------------------
# Kill Excel if it is stuck
# ------------------------------------------------------------

def kill_excel_processes():
    """
    Force-close all Excel processes.

    WARNING:
    This closes every open Excel window on the computer.
    Use only when running the batch and no other Excel files are open.
    """
    subprocess.run(
        ["taskkill", "/F", "/IM", "EXCEL.EXE"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        text=True,
    )


# ------------------------------------------------------------
# Recalculate Excel safely with xlwings
# ------------------------------------------------------------

def recalculate_excel_with_xlwings(
    excel_file,
    max_retries=5,
    wait_seconds=10,
    kill_excel_on_retry=True,
):
    """
    Open an Excel file with xlwings, force recalculation, save, and close.

    This safer version is useful for batch processing because Microsoft Excel
    sometimes fails to open a workbook immediately after it was created,
    especially if the file is inside Google Drive or another synced folder.

    Parameters
    ----------
    excel_file:
        Excel workbook to recalculate.

    max_retries:
        Number of attempts before failing.

    wait_seconds:
        Seconds to wait before opening and between retries.

    kill_excel_on_retry:
        If True, force-closes Excel after a failed attempt.

        WARNING:
        This closes all open Excel windows.
    """
    import xlwings as xw

    excel_file = Path(excel_file).resolve()

    if not excel_file.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    last_error = None

    for attempt in range(1, max_retries + 1):
        app = None
        wb = None

        try:
            print()
            print(f"[Excel recalculation] Attempt {attempt}/{max_retries}")
            print(f"Excel file: {excel_file}")

            # Important when file was just created/saved.
            time.sleep(wait_seconds)

            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            app.screen_updating = False

            wb = app.books.open(str(excel_file))

            # Force recalculation.
            app.calculate()

            # Save and close cleanly.
            wb.save()
            wb.close()
            wb = None

            app.quit()
            app = None

            print(f"Recalculated and saved: {excel_file}")
            return

        except Exception as e:
            last_error = e

            print()
            print("=" * 80)
            print(f"WARNING: Excel recalculation failed on attempt {attempt}/{max_retries}")
            print(f"File: {excel_file}")
            print(f"Error: {e}")
            print("=" * 80)

            # Try to close workbook safely.
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

            # Try to quit Excel safely.
            try:
                if app is not None:
                    app.quit()
            except Exception:
                pass

            # Kill stuck Excel only after a failed attempt.
            if kill_excel_on_retry:
                print("Closing leftover Excel processes...")
                kill_excel_processes()
                time.sleep(3)

            if attempt < max_retries:
                print(f"Waiting {wait_seconds} seconds before retry...")
                time.sleep(wait_seconds)

    raise RuntimeError(
        f"Excel recalculation failed after {max_retries} attempts.\n"
        f"File: {excel_file}\n"
        f"Last error: {last_error}"
    )
