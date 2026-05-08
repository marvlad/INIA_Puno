# excel_builder.py

from pathlib import Path
from copy import copy
import unicodedata
import re
import warnings

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ============================================================
# Allowed crops
# ============================================================

ALLOWED_PRODUCTS = [
    "ACELGA",
    "AGUAYMANTO",
    "AJI",
    "AJO",
    "ALCACHOFA",
    "ALFALFA",
    "ALFALFA +RYE GRASS",
    "ALGODON",
    "APIO",
    "ARROZ",
    "ARVEJA",
    "AVENA + VICIA",
    "BETARRAGA",
    "BROCOLI",
    "CACAO",
    "CAFE",
    "CAIGUA",
    "CALABAZA",
    "CAMOTE",
    "CAMU CAMU",
    "CAÑA DE AZUCAR",
    "CAÑIHUA",
    "CASTAÑA",
    "CEBADA",
    "CEBOLLA",
    "CEBOLLA CHINA",
    "CENTENO",
    "CHÍA",
    "CHIRIMOYA",
    "CIRUELO",
    "CITRICOS",
    "COCO",
    "COCOTERO",
    "COL",
    "COLIFLOR",
    "COPOAZÚ",
    "CULANTRO",
    "DURAZNO",
    "ESPARRAGO",
    "ESPINACA",
    "FLORES",
    "FORESTALES",
    "FRESA",
    "FRIJOL",
    "GARBANZO",
    "GIRASOL",
    "GRANADILLA",
    "GUANÁBANA",
    "GUAYABO",
    "HABA",
    "HIGUERA",
    "HORTALIZAS",
    "LECHUGA",
    "LIMÓN",
    "LLACÓN",
    "MACA",
    "MAIZ AMILACEO (GRANO)",
    "MAIZ MORADO",
    "MANGO",
    "MANI",
    "MANZANO Y PERAL",
    "MELON",
    "NA FORRAJERA",
    "NABO",
    "OCA",
    "OLIVO",
    "OLLUCO",
    "PACAE",
    "PALMA ACEITERA",
    "PALTO",
    "PAN DE ÁRBOL",
    "PAPA MEJORADA",
    "PAPA NATIVA",
    "PAPAYA",
    "PAPAYITA",
    "PASTO NATURAL",
    "PASTOS ASOCIADOS",
    "PEPINO",
    "PIÑA",
    "PITAHAYA",
    "PLATANO",
    "QUINUA",
    "RABANO",
    "ROCOTO",
    "RYE GRASS",
    "SÁBILA",
    "SANDIA",
    "SOYA",
    "TABACO",
    "TARA",
    "TARWI",
    "TOMATE",
    "TREBOL",
    "TRIGO",
    "TUMBO",
    "TUNA",
    "VID",
    "YUCA",
    "ZANAHORIA",
]


# ============================================================
# Config
# ============================================================

TARGET_SHEET = "Base_Datos"
TARGET_HEADER_ROW = 2
TARGET_ROW_TO_REPLACE = 3

IMAGE_SCALE = 0.40


# ============================================================
# Helpers
# ============================================================

def norm(text):
    if text is None:
        return ""

    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("\n", " ")
    return text.strip()


def to_float(value):
    if value is None:
        return None

    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return None


def find_header_row(ws, required_header="NOMBRES Y APELLIDOS", max_rows=30):
    required = norm(required_header)

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, values_only=True),
        start=1,
    ):
        normalized_row = [norm(cell) for cell in row]

        if required in normalized_row:
            return row_index, row

    raise ValueError(f"Could not find header row with: {required_header}")


def get_headers_from_values(header_row_values):
    headers = {}

    for index, value in enumerate(header_row_values, start=1):
        if value is not None:
            key = norm(value)
            headers.setdefault(key, []).append(index)

    return headers


def get_headers_from_sheet(ws, header_row):
    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value

        if value is not None:
            key = norm(value)
            headers.setdefault(key, []).append(col)

    return headers


def get(row_values, headers, header_name):
    cols = headers.get(norm(header_name))

    if cols is None:
        return None

    for col in cols:
        if col - 1 < len(row_values):
            value = row_values[col - 1]

            if value not in (None, ""):
                return value

    return None


def find_person_row(ws, headers, name, start_row):
    name_cols = headers.get(norm("NOMBRES Y APELLIDOS"))

    if name_cols is None:
        raise ValueError("Column 'NOMBRES Y APELLIDOS' was not found.")

    name_col = name_cols[0]
    possible_matches = []

    for row_index, row in enumerate(
        ws.iter_rows(min_row=start_row + 1, values_only=True),
        start=start_row + 1,
    ):
        cell_value = row[name_col - 1] if name_col - 1 < len(row) else None

        if norm(cell_value) == norm(name):
            return row_index, row

        if norm(name) in norm(cell_value) or norm(cell_value) in norm(name):
            possible_matches.append((row_index, cell_value))

    print(f"\nPersona no encontrada exactamente: {name}")

    if possible_matches:
        print("\nPosibles coincidencias:")
        for row_index, value in possible_matches:
            print(f"  Row {row_index}: {value}")

    raise ValueError("Person not found.")


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if src.has_style:
            dst._style = copy(src._style)

        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def set_target_value(ws, target_headers, target_header, value, target_row):
    cols = target_headers.get(norm(target_header))

    if cols is None:
        print(f"Target header not found: {target_header}")
        return False

    col = cols[0]
    ws.cell(target_row, col).value = value
    return True


def validate_crop(cultivo):
    allowed_lookup = {norm(product): product for product in ALLOWED_PRODUCTS}
    cultivo_key = norm(cultivo)

    if cultivo_key not in allowed_lookup:
        print(f'\nCultivo no disponible: "{cultivo}"')
        print("\nCultivos disponibles:")

        for product in ALLOWED_PRODUCTS:
            print(f"  - {product}")

        raise ValueError(f'Cultivo no disponible: "{cultivo}"')

    return allowed_lookup[cultivo_key]


# ============================================================
# Images
# ============================================================

def reinsert_images(wb, image_dir):
    """
    Reinsert images that openpyxl may remove when saving.

    Expected folder:
        extracted_images/

    Expected files:
        image2.png
        image3.jpeg
        image4.png
        image5.jpeg
    """

    image_dir = Path(image_dir)

    if not image_dir.exists():
        print(f"Image directory not found. Skipping image insertion: {image_dir}")
        return

    images_by_sheet = {
        "Textura": [
            ("image2.png", "F1"),
        ],

        "Interpretación": [
            ("image3.jpeg", "A1"),
            ("image4.png", "AF1"),
            ("image5.jpeg", "R1"),
        ],

        "Gráfico_Int": [
            ("image3.jpeg", "A1"),
            ("image4.png", "AD1"),
            ("image5.jpeg", "T1"),
        ],

        "Nec_fert": [
            ("image3.jpeg", "A1"),
            ("image5.jpeg", "E1"),
            ("image4.png", "I1"),
        ],

        "Rec_fert": [
            ("image3.jpeg", "A1"),
            ("image5.jpeg", "D1"),
            ("image4.png", "H1"),
        ],
    }

    # Also support incorrectly encoded sheet names if they exist.
    fallback_sheet_names = {
        "Interpretación": ["Interpretación", "InterpretaciÃ³n"],
        "Gráfico_Int": ["Gráfico_Int", "GrÃ¡fico_Int"],
    }

    for sheet_name, images in images_by_sheet.items():
        actual_sheet_name = None

        candidate_names = fallback_sheet_names.get(sheet_name, [sheet_name])

        for candidate in candidate_names:
            if candidate in wb.sheetnames:
                actual_sheet_name = candidate
                break

        if actual_sheet_name is None:
            print(f"Sheet not found for image insertion: {sheet_name}")
            continue

        ws = wb[actual_sheet_name]

        # Prevent duplicate images if openpyxl preserved some.
        ws._images = []

        for image_filename, anchor in images:
            image_path = image_dir / image_filename

            if not image_path.exists():
                print(f"Image file not found: {image_path}")
                continue

            img = Image(str(image_path))
            img.width = img.width * IMAGE_SCALE
            img.height = img.height * IMAGE_SCALE

            ws.add_image(img, anchor)

            print(
                f"Inserted {image_filename} into {actual_sheet_name}!{anchor} "
                f"at {int(IMAGE_SCALE * 100)}% size"
            )


# ============================================================
# Main builder
# ============================================================

def build_excel_from_template(
    resultados_excel,
    template_excel,
    name,
    cultivo,
    output_excel,
    image_dir="extracted_images",
):
    """
    Creates a filled Excel file from:

        RESULTADOS USUARIOS 2M_Illpa_2.0.xlsx
        Software_Mejorado_Cultivos_Anuales_2025-2026_Arapa.xlsx

    It finds the person by NOMBRES Y APELLIDOS,
    fills Base_Datos row 3 in the template,
    sets the crop,
    chooses Olsen/Bray based on pH,
    and saves output_excel.
    """

    resultados_excel = Path(resultados_excel).resolve()
    template_excel = Path(template_excel).resolve()
    output_excel = Path(output_excel).resolve()
    image_dir = Path(image_dir).resolve()

    if not resultados_excel.exists():
        raise FileNotFoundError(f"RESULTADOS Excel not found: {resultados_excel}")

    if not template_excel.exists():
        raise FileNotFoundError(f"Template Excel not found: {template_excel}")

    output_excel.parent.mkdir(parents=True, exist_ok=True)

    user_product = validate_crop(cultivo)

    print("\nBuilding Excel from template")
    print(f"RESULTADOS Excel: {resultados_excel}")
    print(f"Template Excel: {template_excel}")
    print(f"Person: {name}")
    print(f"Cultivo: {user_product}")
    print(f"Output Excel: {output_excel}")

    # ------------------------------------------------------------
    # 1. Read person data from RESULTADOS
    # ------------------------------------------------------------

    input_wb = load_workbook(resultados_excel, data_only=True, read_only=True)
    input_ws = input_wb.active

    input_header_row_number, input_header_row_values = find_header_row(input_ws)
    input_headers = get_headers_from_values(input_header_row_values)

    person_row_number, person_row_values = find_person_row(
        input_ws,
        input_headers,
        name,
        input_header_row_number,
    )

    input_cultivo = get(person_row_values, input_headers, "CULTIVO A INSTALAR")
    ph_value = get(person_row_values, input_headers, "pH")
    p_value = get(person_row_values, input_headers, "P_mg/kg)")

    ph_number = to_float(ph_value)
    p_number = to_float(p_value)

    codigo = get(person_row_values, input_headers, "CODIGO")

    print("\nFound person:")
    print(f"  Name: {name}")
    print(f"  Input Excel row: {person_row_number}")
    print(f"  CODIGO: {codigo}")
    print(f"  CULTIVO A INSTALAR from input file: {input_cultivo}")
    print(f"  Plan de Recomendación from user: {user_product}")
    print(f"  pH: {ph_value}")
    print(f"  P_mg/kg): {p_value}")

    if ph_number is None:
        print("  Phosphorus method: pH invalid or empty")
    elif ph_number < 7.0:
        print("  Phosphorus method: Bray because pH < 7.0")
    else:
        print("  Phosphorus method: Olsen because pH >= 7.0")

    # ------------------------------------------------------------
    # 2. Open template
    # ------------------------------------------------------------

    target_wb = load_workbook(template_excel)

    if TARGET_SHEET not in target_wb.sheetnames:
        print("Available sheets:")
        for sheet in target_wb.sheetnames:
            print(repr(sheet))

        raise KeyError(f"Target sheet not found: {TARGET_SHEET}")

    target_ws = target_wb[TARGET_SHEET]
    target_headers = get_headers_from_sheet(target_ws, TARGET_HEADER_ROW)

    target_row_number = TARGET_ROW_TO_REPLACE

    copy_row_style(
        target_ws,
        TARGET_ROW_TO_REPLACE,
        target_row_number,
    )

    print(f"\nWriting to target sheet: {TARGET_SHEET}")
    print(f"Replacing target Excel row: {target_row_number}")

    # ------------------------------------------------------------
    # 3. Mapping: target header -> input header
    # ------------------------------------------------------------

    mapping = {
        "CÓDIGO": "CODIGO",
        "Código Laboratorio": "CODIGO",
        "Cotizacion Servicio": "Nº DE COTIZACION",
        "Fecha Muestreo": "FECHA DE MUESTREO",
        "Hora Muestreo": "HORA DE MUESTREO",

        "Codigo_Muestra_Cliente": "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
        "Cliente": "INSTITUCION U ORGANIZACION",
        "Propietario / Productor": "NOMBRES Y APELLIDOS",
        "Direccion del Cliente": "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
        "Solicitado por": "RESPONSABLE",
        "Muestreado por": "NOMBRES Y APELLIDOS",

        "Procedencia de la Muestra": "DIST",

        "Acidez Intercambiable": "Acidez (H+) cmol(+)/Kg",
        "Aluminio Intercambiable": "Aluminio Intercambiable cmol(+)/Kg",

        "Ca_Cmol/Kg": "Calcio (Ca) (*)cmol(+)/Kg",
        "Mg_Cmol/Kg": "Magnesio (Mg) (*) cmol(+)/Kg",
        "Na_Cmol/Kg": "Sodio (Na) (*)cmol(+)/Kg",
        "K_Cmol/Kg": "Potasio (K) (*) cmol(+)/Kg",

        "Calcio Intercambiable": "Calcio (Ca) (*)cmol(+)/Kg",
        "Magnesio Intercambiable": "Magnesio (Mg) (*) cmol(+)/Kg",
        "Sodio Intercambiable": "Sodio (Na) (*)cmol(+)/Kg",
        "Potasio Intercambiable": "Potasio (K) (*) cmol(+)/Kg",

        "Carbonato de Calcio Equivalente": "CaCO3 _% Equivalente",
        "Materia Orgánica (AS-07 Método de Walkley y Black)": "MO_%",
        "Materia Organica (AS-07 Walkley y Black)": "MO_%",
        "Materia Orgánica por LECO": "MO_%",

        "Conductividad Electrica (Suelo)": "CE_mS/m",
        "pH. (Suelo)": "pH",

        "Fósforo Disponible (Bray y Kurtz)": "P_mg/kg)",
        "Fósforo Disponible Bray mpaes": "P_mg/kg)",
        "Potasio Disponible (MPAES)": "K_ppm",
        "Potasio Disponible (AA)": "K_ppm",

        "Arena": "Arena",
        "Arcilla": "Arcilla",
        "Limo": "Limo",
        "Clase Textural": "Clase Textural",

        "Nitrógeno Total": "N_%",
        "Nitrógeno Total Kjeldahl": "N_%",
    }

    # ------------------------------------------------------------
    # 4. Write mapped values
    # ------------------------------------------------------------

    print("\nWriting mapped values:")

    for target_header, input_header in mapping.items():
        value = get(person_row_values, input_headers, input_header)

        if value is None:
            value = 0
            print(f"  {target_header}: INPUT HEADER NOT FOUND or EMPTY -> set to 0")
        else:
            print(f"  {target_header}: {value}")

        set_target_value(
            target_ws,
            target_headers,
            target_header,
            value,
            target_row_number,
        )

    # ------------------------------------------------------------
    # 5. Fixed values
    # ------------------------------------------------------------

    fixed_values = {
        "Número de Muestras": 1,
        "Presentación Muestra": "1 kg",
        "Plan de Recomendación de Fertilización": user_product,
    }

    print("\nWriting fixed values:")

    for target_header, value in fixed_values.items():
        ok = set_target_value(
            target_ws,
            target_headers,
            target_header,
            value,
            target_row_number,
        )

        if ok:
            print(f"  {target_header}: {value}")

    # ------------------------------------------------------------
    # 6. Force specific cells
    # ------------------------------------------------------------

    # T3 = Plan de Recomendación de Fertilización
    target_ws["T3"] = user_product
    print(f"\nT3 / Plan de Recomendación de Fertilización: {user_product}")

    # Phosphorus method:
    # AL3 = Fósforo Disponible Olsen
    # AM3 = Fósforo Disponible Bray y Kurtz
    target_ws["AL3"] = 0
    target_ws["AM3"] = 0

    if p_number is None:
        print("P_mg/kg) not found or invalid. AL3 and AM3 set to 0.")

    elif ph_number is None:
        print("pH not found or invalid. Cannot choose Olsen/Bray. AL3 and AM3 set to 0.")

    elif ph_number < 7.0:
        target_ws["AM3"] = p_number

        print(f"pH: {ph_number}")
        print("Using Bray because pH < 7.0")
        print("AL3 / Fósforo Disponible Olsen: 0")
        print(f"AM3 / Fósforo Disponible Bray y Kurtz: {p_number}")

    else:
        target_ws["AL3"] = p_number

        print(f"pH: {ph_number}")
        print("Using Olsen because pH >= 7.0")
        print(f"AL3 / Fósforo Disponible Olsen: {p_number}")
        print("AM3 / Fósforo Disponible Bray y Kurtz: 0")

    # ------------------------------------------------------------
    # 7. Reinsert images
    # ------------------------------------------------------------

    reinsert_images(
        target_wb,
        image_dir=image_dir,
    )

    # ------------------------------------------------------------
    # 8. Save
    # ------------------------------------------------------------

    target_wb.save(output_excel)

    print("\nExcel created successfully:")
    print(f"  {output_excel}")

    return output_excel
