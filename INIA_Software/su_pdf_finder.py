# su_pdf_finder.py

import os
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from utils import copy_file_to_dir


def norm(text):
    if text is None:
        return ""

    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("\n", " ")
    return text.strip()


def extract_su_number_from_text(text):
    """
    Extracts 723 from strings like:

        SU723-ILL-24
        SU0723-ILL-24
        SU 723
        SU723
        SU723-724
    """

    if not text:
        return None

    m = re.search(r"SU\s*0*(\d+)", str(text), re.IGNORECASE)

    if not m:
        return None

    return int(m.group(1))


def extract_year_from_text(text):
    """
    Extracts year-like values from text.

    Examples:
        SU723-ILL-24      -> 24
        N° 010004-26 ...  -> 26
        2026              -> 26
    """

    if not text:
        return None

    text = str(text)

    # Full year: 2024, 2025, 2026, etc.
    m = re.search(r"\b20(\d{2})\b", text)
    if m:
        return int(m.group(1))

    # Two-digit year after dash, common in filenames/codes
    matches = re.findall(r"[-_ ](\d{2})(?:\D|$)", text)
    if matches:
        return int(matches[-1])

    return None


def find_header_row(ws, required_header="NOMBRES Y APELLIDOS", max_rows=30):
    required = norm(required_header)

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, values_only=True),
        start=1,
    ):
        normalized_row = [norm(cell) for cell in row]

        if required in normalized_row:
            return row_index, row

    raise ValueError(f"Could not find header row with: {required_header}")


def get_headers_from_values(header_row_values):
    headers = {}

    for index, value in enumerate(header_row_values, start=1):
        if value is not None:
            key = norm(value)
            headers.setdefault(key, []).append(index)

    return headers


def find_person_row(ws, headers, name, start_row):
    name_cols = headers.get(norm("NOMBRES Y APELLIDOS"))

    if name_cols is None:
        raise ValueError("Column 'NOMBRES Y APELLIDOS' was not found.")

    name_col = name_cols[0]
    possible_matches = []

    for row_index, row in enumerate(
        ws.iter_rows(min_row=start_row + 1, values_only=True),
        start=start_row + 1,
    ):
        cell_value = row[name_col - 1] if name_col - 1 < len(row) else None

        if norm(cell_value) == norm(name):
            return row_index, row

        if norm(name) in norm(cell_value) or norm(cell_value) in norm(name):
            possible_matches.append((row_index, cell_value))

    print(f"\nPersona no encontrada exactamente: {name}")

    if possible_matches:
        print("\nPosibles coincidencias:")
        for row_index, value in possible_matches:
            print(f"  Row {row_index}: {value}")

    raise ValueError("Person not found.")


def get_value_from_row(row_values, headers, header_name):
    cols = headers.get(norm(header_name))

    if cols is None:
        return None

    for col in cols:
        if col - 1 < len(row_values):
            value = row_values[col - 1]

            if value not in (None, ""):
                return value

    return None


def get_first_available_value(row_values, headers, possible_header_names):
    """
    Tries several possible column names and returns the first value found.
    Useful because sometimes your Excel may have ZONA, LABORATORIO, DEP, etc.
    """

    for header_name in possible_header_names:
        value = get_value_from_row(row_values, headers, header_name)
        if value not in (None, ""):
            return value

    return None


def get_su_info_from_resultados_excel(resultados_excel, person_name):
    """
    Opens RESULTADOS USUARIOS 2M_Illpa_2.0.xlsx,
    finds the row for person_name,
    reads CODIGO and extra metadata,
    and extracts the SU number.

    Returns:
        {
            "su_number": 723,
            "codigo": "SU723-ILL-24",
            "year": 24,
            "place": "PUNO",
            "lab": "ILLPA"
        }
    """

    resultados_excel = Path(resultados_excel).resolve()

    if not resultados_excel.exists():
        raise FileNotFoundError(f"RESULTADOS Excel file not found: {resultados_excel}")

    wb = load_workbook(resultados_excel, data_only=True, read_only=True)
    ws = wb.active

    header_row_number, header_row_values = find_header_row(ws)
    headers = get_headers_from_values(header_row_values)

    person_row_number, person_row_values = find_person_row(
        ws,
        headers,
        person_name,
        header_row_number,
    )

    codigo = get_value_from_row(person_row_values, headers, "CODIGO")
    su_number = extract_su_number_from_text(codigo)

    if su_number is None:
        raise ValueError(f"Could not extract SU number from CODIGO value: {codigo}")

    year = extract_year_from_text(codigo)

    place = get_first_available_value(
        person_row_values,
        headers,
        [
            "ZONA",
            "DEP",
            "DEPARTAMENTO",
            "PROV",
            "PROVINCIA",
            "DIST",
            "DISTRITO",
            "LOCALIDAD",
            "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
        ],
    )

    lab = get_first_available_value(
        person_row_values,
        headers,
        [
            "LABORATORIO",
            "CODIGO LABORATORIO",
            "Código Laboratorio",
        ],
    )

    info = {
        "su_number": su_number,
        "codigo": codigo,
        "year": year,
        "place": place,
        "lab": lab,
        "person_row_number": person_row_number,
    }

    print("\n[6] SU information extracted from RESULTADOS Excel")
    print(f"Person row: {person_row_number}")
    print(f"CODIGO: {codigo}")
    print(f"SU number: {su_number}")

    if year is not None:
        print(f"Year detected: {year}")

    if place:
        print(f"Place detected: {place}")

    if lab:
        print(f"Lab detected: {lab}")

    return info


def get_su_number_from_resultados_excel(resultados_excel, person_name):
    """
    Backward-compatible function.
    If other parts of your code already call this, it will still work.
    """

    info = get_su_info_from_resultados_excel(resultados_excel, person_name)
    return info["su_number"], info["codigo"]


def filename_contains_token(filename, token):
    if not token:
        return False

    filename_n = norm(filename)
    token_n = norm(token)

    return token_n in filename_n


def score_pdf_candidate(pdf_path, su_info):
    """
    Scores a PDF candidate.

    Higher score = better match.

    SU match is already required before this function.
    This function adds points for:
        - year match
        - place match
        - lab match
        - codigo fragments
    """

    filename = pdf_path.name
    filename_n = norm(filename)

    score = 0
    reasons = []

    codigo = su_info.get("codigo")
    year = su_info.get("year")
    place = su_info.get("place")
    lab = su_info.get("lab")

    # Year match, e.g. 24, 25, 26
    filename_year = extract_year_from_text(filename)

    if year is not None and filename_year is not None:
        if year == filename_year:
            score += 50
            reasons.append(f"year match: {year}")
        else:
            score -= 20
            reasons.append(f"year mismatch: Excel={year}, PDF={filename_year}")

    # Place match, e.g. PUNO, CUSCO
    if place and filename_contains_token(filename, place):
        score += 40
        reasons.append(f"place match: {place}")

    # Lab match, e.g. ILLPA
    if lab and filename_contains_token(filename, lab):
        score += 30
        reasons.append(f"lab match: {lab}")

    # Codigo fragments match
    # Example: SU723-ILL-24 -> tokens SU723, ILL, 24
    if codigo:
        codigo_n = norm(codigo)
        tokens = re.split(r"[-_ ]+", codigo_n)

        for token in tokens:
            if len(token) >= 3 and token in filename_n:
                score += 10
                reasons.append(f"codigo token match: {token}")

    return score, reasons


def find_su_pdfs(su_number, folder_path):
    folder_path = Path(folder_path).resolve()

    if not folder_path.exists():
        raise FileNotFoundError(f"PDF folder not found: {folder_path}")

    range_pattern = re.compile(
        r"SU\s*0*(\d+)\s*(?:-\s*0*(\d+))?",
        re.IGNORECASE,
    )

    matches = []

    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            if not filename.lower().endswith(".pdf"):
                continue

            m = range_pattern.search(filename)

            if not m:
                continue

            start = int(m.group(1))
            end = int(m.group(2)) if m.group(2) else start

            if start <= su_number <= end:
                matches.append(Path(root) / filename)

    return matches


def find_best_su_pdf(su_info, folder_path):
    """
    Finds all PDFs matching the SU number,
    then ranks them using year/place/lab/codigo information.
    """

    su_number = su_info["su_number"]
    matches = find_su_pdfs(su_number, folder_path)

    if not matches:
        return None, []

    scored = []

    for pdf in matches:
        score, reasons = score_pdf_candidate(pdf, su_info)
        scored.append(
            {
                "path": pdf,
                "score": score,
                "reasons": reasons,
            }
        )

    scored.sort(key=lambda item: item["score"], reverse=True)

    best = scored[0]

    return best, scored


def copy_su_pdfs_to_report_dir(su_number_or_info, pdf_folder, report_dir, copy_all_matches=False):
    """
    Copies the best original SU PDF file to report_dir.

    You can call this in two ways:

    Old way:
        copy_su_pdfs_to_report_dir(su_number, pdf_folder, report_dir)

    New better way:
        su_info = get_su_info_from_resultados_excel(resultados_excel, person_name)
        copy_su_pdfs_to_report_dir(su_info, pdf_folder, report_dir)

    If copy_all_matches=True, it copies all SU matches.
    """

    print("\n[7] Searching and copying original SU PDF files")

    if su_number_or_info is None:
        print("No SU number provided or detected. Skipping original PDF copy.")
        return []

    if isinstance(su_number_or_info, dict):
        su_info = su_number_or_info
        su_number = su_info["su_number"]
    else:
        su_number = su_number_or_info
        su_info = {
            "su_number": su_number,
            "codigo": None,
            "year": None,
            "place": None,
            "lab": None,
        }

    print(f"SU number: SU{su_number:04d}")
    print(f"Search folder: {pdf_folder}")

    best, scored = find_best_su_pdf(su_info, pdf_folder)

    if not scored:
        print(f"No PDF found for SU{su_number:04d}")
        return []

    print("\nPDF candidates found:")

    for item in scored:
        pdf = item["path"]
        score = item["score"]
        reasons = ", ".join(item["reasons"]) if item["reasons"] else "SU match only"

        print(f"  Score {score:>4}: {pdf.name}")
        print(f"          {reasons}")

    copied = []

    if copy_all_matches:
        print("\nCopying all matching PDFs.")

        for item in scored:
            pdf = item["path"]
            dst = copy_file_to_dir(pdf, report_dir)

            if dst:
                copied.append(dst)
                print(f"Copied: {pdf}")
                print(f"    to: {dst}")

        return copied

    best_pdf = best["path"]
    best_score = best["score"]

    # Optional ambiguity warning
    if len(scored) > 1:
        second_score = scored[1]["score"]

        if best_score == second_score:
            print("\nWARNING: Multiple PDFs have the same best score.")
            print("The first one will be copied, but you should review the candidates.")

    dst = copy_file_to_dir(best_pdf, report_dir)

    if dst:
        copied.append(dst)
        print("\nBest PDF copied:")
        print(f"Copied: {best_pdf}")
        print(f"    to: {dst}")

    return copied
