# filter_by_distrito.py

from pathlib import Path
import argparse
import unicodedata
import re
import csv

from openpyxl import load_workbook


def norm(text):
    """
    Normalize text for comparison:
    - removes accents
    - uppercases
    - removes extra spaces
    """
    if text is None:
        return ""

    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def find_header_row(ws, required_headers=None, max_rows=30):
    """
    Finds the header row by looking for key headers.
    """
    if required_headers is None:
        required_headers = [
            "DIST",
            "NOMBRES Y APELLIDOS",
            "CULTIVO A INSTALAR",
        ]

    required_norm = [norm(h) for h in required_headers]

    for row_number in range(1, max_rows + 1):
        row_values = [
            ws.cell(row_number, col).value
            for col in range(1, ws.max_column + 1)
        ]

        normalized_row = [norm(value) for value in row_values]

        if all(header in normalized_row for header in required_norm):
            return row_number

    raise ValueError(
        "Could not find header row with required headers: "
        + ", ".join(required_headers)
    )


def get_headers(ws, header_row):
    """
    Returns dictionary:
        normalized header name -> column number
    """
    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value

        if value is not None:
            headers[norm(value)] = col

    return headers


def filter_by_distrito(input_excel, distrito, output_csv, sheet_name=None):
    input_excel = Path(input_excel)
    output_csv = Path(output_csv)

    if not input_excel.exists():
        raise FileNotFoundError(f"Input Excel file not found: {input_excel}")

    wb = load_workbook(input_excel, data_only=True, read_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        ws = wb[sheet_name]
    else:
        ws = wb.active

    header_row = find_header_row(ws)
    headers = get_headers(ws, header_row)

    dist_col = headers[norm("DIST")]
    name_col = headers[norm("NOMBRES Y APELLIDOS")]
    cultivo_col = headers[norm("CULTIVO A INSTALAR")]

    distrito_target = norm(distrito)

    results = []

    for row in range(header_row + 1, ws.max_row + 1):
        dist_value = ws.cell(row, dist_col).value

        if norm(dist_value) != distrito_target:
            continue

        name_value = ws.cell(row, name_col).value
        cultivo_value = ws.cell(row, cultivo_col).value

        if name_value in (None, "") and cultivo_value in (None, ""):
            continue

        results.append({
            "NOMBRES Y APELLIDOS": name_value,
            "CULTIVO A INSTALAR": cultivo_value,
        })

    output_csv.parent.mkdir(parents=True, exist_ok=True)

    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "NOMBRES Y APELLIDOS",
                "CULTIVO A INSTALAR",
            ],
        )

        writer.writeheader()
        writer.writerows(results)

    print()
    print(f"Input Excel: {input_excel}")
    print(f"Sheet: {ws.title}")
    print(f"DIST searched: {distrito}")
    print(f"Matches found: {len(results)}")
    print(f"Output CSV: {output_csv}")


def main():
    parser = argparse.ArgumentParser(
        description="Filter Excel rows by DIST and export names/crops to CSV."
    )

    parser.add_argument(
        "--input-excel",
        required=True,
        help="Input Excel file.",
    )

    parser.add_argument(
        "--dist",
        required=True,
        help="Distrito to search in the DIST column.",
    )

    parser.add_argument(
        "--output-csv",
        default="filtered_by_dist.csv",
        help="Output CSV file.",
    )

    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional sheet name. If not provided, the active sheet is used.",
    )

    args = parser.parse_args()

    filter_by_distrito(
        input_excel=args.input_excel,
        distrito=args.dist,
        output_csv=args.output_csv,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()
