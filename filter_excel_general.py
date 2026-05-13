# filter_excel_general.py

from pathlib import Path
import argparse
import unicodedata
import re
import csv

from openpyxl import load_workbook


# ------------------------------------------------------------
# Expected INIA columns
# ------------------------------------------------------------

ALL_COLUMNS = [
    "Nº",
    "DEP",
    "PROV",
    "DIST",
    "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
    "NOMBRES Y APELLIDOS",
    "DNI",
    "TELÉFONO",
    "FECHA DE MUESTREO",
    "HORA DE MUESTREO",
    "CULTIVO ANTERIOR",
    "VARIEDAD (ANTERIOR)",
    "CULTIVO A INSTALAR",
    "VARIEDAD (A INSTALAR)",
    "HECTÁREAS",
    "COORDENADAS (E)",
    "COORDENADAS (N)",
    "ALTITUD (m.s.n.m.)",
    "CODIGO",
    "LABORATORIO",
    "ZONA",
    "ESTADO",
    "Nº DE COTIZACION",
    "INSTITUCION U ORGANIZACION",
    "RESPONSABLE",
    "CORREO/CEL",
    "OBS",
]


# Output columns required by run_main_from_csv_parallel.py
DEFAULT_OUTPUT_COLUMNS = [
    "NOMBRES Y APELLIDOS",
    "CULTIVO A INSTALAR",
]


# ------------------------------------------------------------
# Text helpers
# ------------------------------------------------------------

def norm(text):
    """
    Normalize text for comparison:
    - removes accents
    - uppercases
    - removes extra spaces
    """
    if text is None:
        return ""

    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def safe_filename_text(text):
    """
    Convert text into safe filename text.
    """
    if text is None:
        return "unknown"

    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"\s+", "_", text)

    return text.strip("_") or "unknown"


def make_output_name_from_filters(filters):
    """
    Create CSV filename from one or multiple filters.

    Example:
        DIST=AYAVIRI

    Output:
        DIST_AYAVIRI.csv

    Example:
        PROV=MELGAR
        DIST=AYAVIRI
        CULTIVO A INSTALAR=ALFALFA

    Output:
        PROV_MELGAR__DIST_AYAVIRI__CULTIVO_A_INSTALAR_ALFALFA.csv
    """
    parts = []

    for item in filters:
        column = safe_filename_text(item["column"])
        value = safe_filename_text(item["value"])
        parts.append(f"{column}_{value}")

    filename = "__".join(parts)

    if len(filename) > 180:
        filename = filename[:180].rstrip("_")

    return filename + ".csv"


# ------------------------------------------------------------
# Excel helpers
# ------------------------------------------------------------

def find_header_row(ws, max_rows=50):
    """
    Finds the most likely header row using the expected INIA headers.
    """
    known_norm = [norm(h) for h in ALL_COLUMNS]

    max_row = ws.max_row or max_rows
    max_col = ws.max_column or 100

    best_row = None
    best_score = 0

    for row_number in range(1, min(max_row, max_rows) + 1):
        row_values = [
            ws.cell(row_number, col).value
            for col in range(1, max_col + 1)
        ]

        normalized_row = [
            norm(value)
            for value in row_values
            if value is not None and str(value).strip()
        ]

        score = 0
        for header in known_norm:
            if header in normalized_row:
                score += 1

        if score > best_score:
            best_score = score
            best_row = row_number

    if best_row is not None and best_score >= 3:
        return best_row

    raise ValueError(
        f"Could not detect header row in sheet '{ws.title}'. "
        "Please check the Excel format."
    )


def get_headers(ws, header_row):
    """
    Returns:
        headers_norm_to_col: normalized header -> column number
        headers_original: original header names in order
    """
    headers_norm_to_col = {}
    headers_original = []

    max_col = ws.max_column or 100

    for col in range(1, max_col + 1):
        value = ws.cell(header_row, col).value

        if value is not None and str(value).strip():
            original = str(value).strip()
            normalized = norm(original)

            headers_norm_to_col[normalized] = col
            headers_original.append(original)

    return headers_norm_to_col, headers_original


def find_data_sheet(wb, sheet_name=None):
    """
    If sheet_name is given, use that sheet.
    Otherwise, search all sheets until one looks like a data sheet.
    """
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        ws = wb[sheet_name]
        find_header_row(ws)
        return ws

    for ws in wb.worksheets:
        try:
            find_header_row(ws)
            print(f"Detected data sheet: {ws.title}")
            return ws
        except ValueError:
            continue

    print("Available sheets:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

    raise ValueError("Could not find a suitable data sheet.")


# ------------------------------------------------------------
# Filter helpers
# ------------------------------------------------------------

def parse_filters_from_text(filters_text):
    """
    Parses filters from text.

    Accepted format, one filter per line:

        DIST=AYAVIRI
        PROV=MELGAR
        CULTIVO A INSTALAR=ALFALFA

    Also accepts ':' instead of '=':

        DIST: AYAVIRI
        PROV: MELGAR
    """
    filters = []

    if filters_text is None:
        return filters

    lines = str(filters_text).splitlines()

    for line in lines:
        line = line.strip()

        if not line:
            continue

        if "=" in line:
            column, value = line.split("=", 1)
        elif ":" in line:
            column, value = line.split(":", 1)
        else:
            raise ValueError(
                f"Invalid filter line: {line}\n"
                "Use format COLUMN=VALUE, for example: DIST=AYAVIRI"
            )

        column = column.strip()
        value = value.strip()

        if not column or not value:
            raise ValueError(
                f"Invalid filter line: {line}\n"
                "Both column and value are required."
            )

        filters.append({
            "column": column,
            "value": value,
        })

    return filters


def value_matches(cell_value, target_value, mode="exact"):
    """
    mode:
        exact    -> normalized cell == normalized target
        contains -> normalized target inside normalized cell
    """
    cell_norm = norm(cell_value)
    target_norm = norm(target_value)

    if mode == "exact":
        return cell_norm == target_norm

    if mode == "contains":
        return target_norm in cell_norm

    raise ValueError(f"Unknown match mode: {mode}")


def validate_filter_columns(filters, headers_norm_to_col, headers_original):
    """
    Checks that all requested filter columns exist in the Excel file.
    """
    missing = []

    for item in filters:
        if norm(item["column"]) not in headers_norm_to_col:
            missing.append(item["column"])

    if missing:
        raise ValueError(
            "Some filter columns were not found:\n"
            + "\n".join(f"  - {col}" for col in missing)
            + "\n\nAvailable columns:\n"
            + "\n".join(f"  - {h}" for h in headers_original)
        )


def validate_output_columns(headers_norm_to_col, headers_original):
    """
    Checks that the two required output columns exist.
    """
    missing = []

    for col in DEFAULT_OUTPUT_COLUMNS:
        if norm(col) not in headers_norm_to_col:
            missing.append(col)

    if missing:
        raise ValueError(
            "The required output columns were not found:\n"
            + "\n".join(f"  - {col}" for col in missing)
            + "\n\nAvailable columns:\n"
            + "\n".join(f"  - {h}" for h in headers_original)
        )


def row_matches_all_filters(ws, row, filters, headers_norm_to_col, match_mode="exact"):
    """
    Returns True only if the row matches all filters.
    """
    for item in filters:
        column = item["column"]
        value = item["value"]

        column_norm = norm(column)

        if column_norm not in headers_norm_to_col:
            raise ValueError(f"Filter column not found: {column}")

        col_index = headers_norm_to_col[column_norm]
        cell_value = ws.cell(row, col_index).value

        if not value_matches(cell_value, value, mode=match_mode):
            return False

    return True


# ------------------------------------------------------------
# Main filtering function
# ------------------------------------------------------------

def filter_excel_general(
    input_excel,
    filters,
    output_csv=None,
    output_dir="outputs",
    sheet_name=None,
    match_mode="exact",
):
    """
    Filter Excel rows by one or multiple filters.

    The output CSV always contains only:
        NOMBRES Y APELLIDOS
        CULTIVO A INSTALAR
    """
    input_excel = Path(input_excel)
    output_dir = Path(output_dir)

    if not input_excel.exists():
        raise FileNotFoundError(f"Input Excel file not found: {input_excel}")

    if not filters:
        raise ValueError("At least one filter is required.")

    if output_csv is None or not str(output_csv).strip():
        output_csv = output_dir / make_output_name_from_filters(filters)
    else:
        output_csv = Path(output_csv)

    wb = load_workbook(input_excel, data_only=True)

    ws = find_data_sheet(wb, sheet_name=sheet_name)

    header_row = find_header_row(ws)
    headers_norm_to_col, headers_original = get_headers(ws, header_row)

    validate_filter_columns(filters, headers_norm_to_col, headers_original)
    validate_output_columns(headers_norm_to_col, headers_original)

    results = []

    for row in range(header_row + 1, ws.max_row + 1):
        if not row_matches_all_filters(
            ws=ws,
            row=row,
            filters=filters,
            headers_norm_to_col=headers_norm_to_col,
            match_mode=match_mode,
        ):
            continue

        output_row = {}

        for col_name in DEFAULT_OUTPUT_COLUMNS:
            col_index = headers_norm_to_col[norm(col_name)]
            output_row[col_name] = ws.cell(row, col_index).value

        # Skip empty rows
        if all(value in (None, "") for value in output_row.values()):
            continue

        results.append(output_row)

    output_csv.parent.mkdir(parents=True, exist_ok=True)

    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=DEFAULT_OUTPUT_COLUMNS,
        )

        writer.writeheader()
        writer.writerows(results)

    print()
    print(f"Input Excel: {input_excel}")
    print(f"Sheet used: {ws.title}")
    print(f"Header row: {header_row}")
    print("Filters:")
    for item in filters:
        print(f"  {item['column']} = {item['value']}")
    print(f"Match mode: {match_mode}")
    print(f"Matches found: {len(results)}")
    print(f"Output CSV: {output_csv}")

    return {
        "input_excel": str(input_excel),
        "sheet_used": ws.title,
        "header_row": header_row,
        "filters": filters,
        "match_mode": match_mode,
        "matches_found": len(results),
        "output_csv": str(output_csv),
        "output_columns": DEFAULT_OUTPUT_COLUMNS,
    }


# ------------------------------------------------------------
# Command-line interface
# ------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Filter Excel rows by one or multiple columns. "
            "Output CSV always contains NOMBRES Y APELLIDOS and CULTIVO A INSTALAR."
        )
    )

    parser.add_argument(
        "--input-excel",
        required=True,
        help="Input Excel file.",
    )

    parser.add_argument(
        "--filter",
        action="append",
        required=True,
        help=(
            "Filter in COLUMN=VALUE format. "
            "Can be repeated. Example: "
            '--filter "DIST=AYAVIRI" --filter "PROV=MELGAR"'
        ),
    )

    parser.add_argument(
        "--output-csv",
        default=None,
        help=(
            "Output CSV file. "
            "If omitted, the filename is generated automatically from the filters."
        ),
    )

    parser.add_argument(
        "--output-dir",
        default="outputs",
        help="Folder for automatic CSV output name. Default: outputs.",
    )

    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional sheet name. If not provided, all sheets are searched.",
    )

    parser.add_argument(
        "--match-mode",
        choices=["exact", "contains"],
        default="exact",
        help="Use exact or contains matching. Default: exact.",
    )

    args = parser.parse_args()

    filters_text = "\n".join(args.filter)
    filters = parse_filters_from_text(filters_text)

    filter_excel_general(
        input_excel=args.input_excel,
        filters=filters,
        output_csv=args.output_csv,
        output_dir=args.output_dir,
        sheet_name=args.sheet,
        match_mode=args.match_mode,
    )


if __name__ == "__main__":
    main()
