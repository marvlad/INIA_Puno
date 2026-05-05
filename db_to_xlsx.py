from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from copy import copy
from pathlib import Path
import unicodedata
import re
import warnings
import sys
import mysql.connector


BASE_DIR = Path(__file__).resolve().parent

TARGET_FILE = BASE_DIR / "Software_Mejorado_Cultivos Anuales_2025-2026_Arapa.xlsx"
IMAGE_DIR = BASE_DIR / "extracted_images"

TARGET_SHEET = "Base_Datos"
TARGET_HEADER_ROW = 2
TARGET_ROW_TO_REPLACE = 3

IMAGE_SCALE = 0.40

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# =========================
# MYSQL CONFIGURATION
# =========================

MYSQL_HOST = "127.0.0.1"
MYSQL_PORT = 3306
MYSQL_USER = "root"
MYSQL_PASSWORD = "your_password"
MYSQL_DATABASE = "soil_db"
MYSQL_TABLE = "soil_samples"


ALLOWED_PRODUCTS = [
    "ACELGA",
    "AGUAYMANTO",
    "AJI",
    "AJO",
    "ALCACHOFA",
    "ALFALFA",
    "RYE GRASS",
    "ALFALFA +RYE GRASS",
    "ALGODON",
    "APIO",
    "ARROZ",
    "ARVEJA",
    "AVENA + VICIA",
    "BETARRAGA",
    "BROCOLI",
    "CACAO",
    "CAFE",
    "CAIGUA",
    "CALABAZA",
    "CAMOTE",
    "CAÑA DE AZUCAR",
    "CEBADA",
    "CEBOLLA",
    "CENTENO",
    "CIRUELO",
    "CITRICOS",
    "COCOTERO",
    "COL",
    "COLIFLOR",
    "DURAZNO",
    "ESPARRAGO",
    "ESPINACA",
    "FRESA",
    "FRIJOL",
    "GARBANZO",
    "GIRASOL",
    "GUAYABO",
    "HABA",
    "HIGUERA",
    "LECHUGA",
    "MAIZ MORADO",
    "MAIZ AMILACEO (GRANO)",
    "MANGO",
    "MANI",
    "MANZANO Y PERAL",
    "MELON",
    "NABO",
    "OLIVO",
    "PALMA ACEITERA",
    "PALTO",
    "PAPA NATIVA",
    "PAPA MEJORADA",
    "PAPAYA",
    "PASTOS ASOCIADOS",
    "PEPINO",
    "PIÑA",
    "PLATANO",
    "RABANO",
    "SANDIA",
    "SOYA",
    "TABACO",
    "TARA",
    "TOMATE",
    "TREBOL",
    "TRIGO",
    "VID",
    "ZANAHORIA",
]


if len(sys.argv) < 3:
    print('Usage: python3 fill_from_db.py "Full Name" "CROP"')
    print()
    print("Example:")
    print('  python3 fill_from_db.py "Hancco Suca Isaac" "palto"')
    sys.exit(1)

PERSON_NAME = sys.argv[1]
USER_PRODUCT_RAW = sys.argv[2].strip()


def norm(text):
    if text is None:
        return ""

    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("\n", " ")
    return text.strip()


def safe_filename(text):
    text = str(text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^A-Za-z0-9_\-]", "", text)
    return text


def to_float(value):
    if value is None:
        return None

    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return None


# ------------------------------------------------------------
# Original Excel headers -> MySQL column names
# ------------------------------------------------------------

INPUT_TO_DB = {
    "Nº": "nro",
    "DEP": "dep",
    "PROV": "prov",
    "DIST": "dist",
    "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)": "localidad_comunidad_caserio_asociacion_etc",
    "NOMBRES Y APELLIDOS": "nombres_y_apellidos",
    "DNI": "dni",
    "TELÉFONO": "telefono",
    "FECHA DE MUESTREO": "fecha_de_muestreo",
    "HORA DE MUESTREO": "hora_de_muestreo",
    "CULTIVO ANTERIOR": "cultivo_anterior",
    "VARIEDAD (ANTERIOR)": "variedad_anterior",
    "CULTIVO A INSTALAR": "cultivo_a_instalar",
    "VARIEDAD (A INSTALAR)": "variedad_a_instalar",
    "HECTÁREAS": "hectareas",
    "COORDENADAS (E)": "coordenadas_e",
    "COORDENADAS (N)": "coordenadas_n",
    "ALTITUD (m.s.n.m.)": "altitud_m_s_n_m",
    "CODIGO": "codigo",
    "LABORATORIO": "laboratorio",
    "ZONA": "zona",
    "ESTADO": "estado",
    "Nº DE COTIZACION": "nro_de_cotizacion",
    "INSTITUCION U ORGANIZACION": "institucion_u_organizacion",
    "RESPONSABLE": "responsable",
    "CORREO/CEL": "correo_cel",
    "OBS": "obs",
    "pH": "ph",
    "CE_mS/m": "ce_ms_m",
    "CaCO3 _% Equivalente": "caco3_porcentaje_equivalente",
    "Aluminio Intercambiable cmol(+)/Kg": "aluminio_intercambiable_cmol_plus_kg",
    "Acidez (H+) cmol(+)/Kg": "acidez_hplus_cmol_plus_kg",
    "MO_%": "mo_porcentaje",
    "Arena": "arena",
    "Arcilla": "arcilla",
    "Limo": "limo",
    "Clase Textural": "clase_textural",
    "Calcio (Ca) (*)cmol(+)/Kg": "calcio_ca_cmol_plus_kg",
    "Magnesio (Mg) (*) cmol(+)/Kg": "magnesio_mg_cmol_plus_kg",
    "Sodio (Na) (*) cmol(+)/Kg": "sodio_na_cmol_plus_kg",
    "Potasio (K) (*) cmol(+)/Kg": "potasio_k_cmol_plus_kg",
    "N_%": "n_porcentaje",
    "K_ppm": "k_ppm",
    "P_mg/kg)": "p_mg_kg",
    "CICe (*) cmol(+)/Kg": "cice_cmol_plus_kg",
}


def get(row_dict, header_name):
    db_col = INPUT_TO_DB.get(header_name)

    if db_col is None:
        return None

    value = row_dict.get(db_col)

    if value in ("", None):
        return None

    return value


def find_person_from_db(person_name):
    conn = mysql.connector.connect(
        host=MYSQL_HOST,
        port=MYSQL_PORT,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DATABASE,
    )

    cursor = conn.cursor(dictionary=True)

    # First: try exact match using MySQL.
    cursor.execute(
        f"""
        SELECT *
        FROM `{MYSQL_TABLE}`
        WHERE nombres_y_apellidos = %s
        LIMIT 5
        """,
        (person_name,),
    )

    rows = cursor.fetchall()

    # If exact MySQL match fails, search broader and compare normalized names in Python.
    if not rows:
        cursor.execute(
            f"""
            SELECT *
            FROM `{MYSQL_TABLE}`
            WHERE nombres_y_apellidos LIKE %s
            LIMIT 20
            """,
            (f"%{person_name}%",),
        )
        rows = cursor.fetchall()

    # If still empty, retrieve possible partial matches by first/last token.
    if not rows:
        tokens = [t for t in norm(person_name).split() if len(t) >= 3]

        possible_matches = []

        for token in tokens[:3]:
            cursor.execute(
                f"""
                SELECT id, nombres_y_apellidos, dni, cultivo_a_instalar
                FROM `{MYSQL_TABLE}`
                WHERE nombres_y_apellidos LIKE %s
                LIMIT 20
                """,
                (f"%{token}%",),
            )
            possible_matches.extend(cursor.fetchall())

        cursor.close()
        conn.close()

        print(f"\nPersona no encontrada: {person_name}")

        if possible_matches:
            print("\nPosibles coincidencias:")
            seen = set()
            for match in possible_matches:
                name = match.get("nombres_y_apellidos")
                if name in seen:
                    continue
                seen.add(name)
                print(
                    f"  ID {match.get('id')}: {name} | "
                    f"DNI: {match.get('dni')} | "
                    f"Cultivo: {match.get('cultivo_a_instalar')}"
                )

        raise ValueError("Person not found in database.")

    # Prefer exact normalized match.
    exact_matches = [
        row for row in rows
        if norm(row.get("nombres_y_apellidos")) == norm(person_name)
    ]

    if exact_matches:
        selected = exact_matches[0]
    else:
        selected = rows[0]

    cursor.close()
    conn.close()

    return selected


def get_headers_from_sheet(ws, header_row):
    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value

        if value is not None:
            key = norm(value)

            if key not in headers:
                headers[key] = []

            headers[key].append(col)

    return headers


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if src.has_style:
            dst._style = copy(src._style)

        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def set_target_value(ws, target_headers, target_header, value, target_row):
    cols = target_headers.get(norm(target_header))

    if cols is None:
        print(f"Target header not found: {target_header}")
        return False

    col = cols[0]
    ws.cell(target_row, col).value = value
    return True


def reinsert_images(wb):
    images_by_sheet = {
        "Textura": [
            ("image2.png", "F1"),
        ],
        "Interpretación": [
            ("image3.jpeg", "A1"),
            ("image4.png", "AF1"),
            ("image5.jpeg", "R1"),
        ],
        "Gráfico_Int": [
            ("image3.jpeg", "A1"),
            ("image4.png", "AD1"),
            ("image5.jpeg", "T1"),
        ],
        "Nec_fert": [
            ("image3.jpeg", "A1"),
            ("image5.jpeg", "E1"),
            ("image4.png", "I1"),
        ],
        "Rec_fert": [
            ("image3.jpeg", "A1"),
            ("image5.jpeg", "D1"),
            ("image4.png", "H1"),
        ],
    }

    for sheet_name, images in images_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet not found for image insertion: {sheet_name}")
            continue

        ws = wb[sheet_name]
        ws._images = []

        for image_filename, anchor in images:
            image_path = IMAGE_DIR / image_filename

            if not image_path.exists():
                print(f"Image file not found: {image_path}")
                continue

            img = Image(str(image_path))
            img.width = img.width * IMAGE_SCALE
            img.height = img.height * IMAGE_SCALE

            ws.add_image(img, anchor)
            print(
                f"Inserted {image_filename} into {sheet_name}!{anchor} "
                f"at {int(IMAGE_SCALE * 100)}% size"
            )


# ------------------------------------------------------------
# 0. Check user crop
# ------------------------------------------------------------

allowed_lookup = {norm(product): product for product in ALLOWED_PRODUCTS}
user_product_key = norm(USER_PRODUCT_RAW)

if user_product_key not in allowed_lookup:
    print(f'\nCultivo no disponible: "{USER_PRODUCT_RAW}"')
    print("\nCultivos disponibles:")
    for product in ALLOWED_PRODUCTS:
        print(f"  - {product}")
    sys.exit(1)

USER_PRODUCT = allowed_lookup[user_product_key]

OUTPUT_FILE = BASE_DIR / (
    "Software_Mejorado_Cultivos_Anuales_2025-2026_Arapa_"
    f"{safe_filename(PERSON_NAME)}_{safe_filename(USER_PRODUCT)}.xlsx"
)


# ------------------------------------------------------------
# 1. Read person from MySQL database
# ------------------------------------------------------------

person_row = find_person_from_db(PERSON_NAME)

input_cultivo = get(person_row, "CULTIVO A INSTALAR")
ph_value = get(person_row, "pH")
p_value = get(person_row, "P_mg/kg)")

ph_number = to_float(ph_value)
p_number = to_float(p_value)


# ------------------------------------------------------------
# 1.1 Display input values from database
# ------------------------------------------------------------

input_display_headers = [
    "Nº",
    "DEP",
    "PROV",
    "DIST",
    "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
    "NOMBRES Y APELLIDOS",
    "DNI",
    "TELÉFONO",
    "FECHA DE MUESTREO",
    "HORA DE MUESTREO",
    "CULTIVO ANTERIOR",
    "VARIEDAD (ANTERIOR)",
    "CULTIVO A INSTALAR",
    "VARIEDAD (A INSTALAR)",
    "HECTÁREAS",
    "COORDENADAS (E)",
    "COORDENADAS (N)",
    "ALTITUD (m.s.n.m.)",
    "CODIGO",
    "LABORATORIO",
    "ZONA",
    "ESTADO",
    "Nº DE COTIZACION",
    "INSTITUCION U ORGANIZACION",
    "RESPONSABLE",
    "CORREO/CEL",
    "OBS",
    "pH",
    "CE_mS/m",
    "CaCO3 _% Equivalente",
    "Aluminio Intercambiable cmol(+)/Kg",
    "Acidez (H+) cmol(+)/Kg",
    "MO_%",
    "Arena",
    "Arcilla",
    "Limo",
    "Clase Textural",
    "Calcio (Ca) (*)cmol(+)/Kg",
    "Magnesio (Mg) (*) cmol(+)/Kg",
    "Sodio (Na) (*)cmol(+)/Kg",
    "Potasio (K) (*) cmol(+)/Kg",
    "N_%",
    "K_ppm",
    "P_mg/kg)",
    "CICe (*) cmol(+)/Kg",
]

print(f"\nFound person: {PERSON_NAME}")
print(f"Database ID: {person_row.get('id')}")
print("=" * 70)
print("DATABASE VALUES")
print("=" * 70)

for header in input_display_headers:
    value = get(person_row, header)
    print(f"{header}: {value}")

print("=" * 70)
print(f"CULTIVO A INSTALAR from database: {input_cultivo}")
print(f"Plan de Recomendación de Fertilización from user: {USER_PRODUCT}")
print(f"pH from database: {ph_value}")
print(f"P_mg/kg) from database: {p_value}")

if ph_number is None:
    print("Phosphorus method: pH invalid or empty")
elif ph_number < 7.0:
    print("Phosphorus method: Bray because pH < 7.0")
else:
    print("Phosphorus method: Olsen because pH >= 7.0")

print("=" * 70)


# ------------------------------------------------------------
# 2. Open target Excel file
# ------------------------------------------------------------

target_wb = load_workbook(TARGET_FILE)

if TARGET_SHEET not in target_wb.sheetnames:
    print("Available sheets:")
    for sheet in target_wb.sheetnames:
        print(repr(sheet))
    raise KeyError(f"Target sheet not found: {TARGET_SHEET}")

target_ws = target_wb[TARGET_SHEET]
target_headers = get_headers_from_sheet(target_ws, TARGET_HEADER_ROW)

target_row_number = TARGET_ROW_TO_REPLACE

copy_row_style(target_ws, TARGET_ROW_TO_REPLACE, target_row_number)

print(f"Writing to target sheet: {TARGET_SHEET}")
print(f"Replacing target Excel row: {target_row_number}")
print("=" * 70)


# ------------------------------------------------------------
# 3. Mapping: target column -> original input column
# ------------------------------------------------------------

mapping = {
    "CÓDIGO": "CODIGO",
    "Código Laboratorio": "CODIGO",
    "Cotizacion Servicio": "Nº DE COTIZACION",
    "Fecha Muestreo": "FECHA DE MUESTREO",
    "Hora Muestreo": "HORA DE MUESTREO",

    "Codigo_Muestra_Cliente": "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
    "Cliente": "INSTITUCION U ORGANIZACION",
    "Propietario / Productor": "NOMBRES Y APELLIDOS",
    "Direccion del Cliente": "LOCALIDAD (COMUNIDAD, CASERÍO, ASOCIACIÓN, ETC)",
    "Solicitado por": "RESPONSABLE",
    "Muestreado por": "NOMBRES Y APELLIDOS",

    "Procedencia de la Muestra": "DIST",

    "Acidez Intercambiable": "Acidez (H+) cmol(+)/Kg",
    "Aluminio Intercambiable": "Aluminio Intercambiable cmol(+)/Kg",

    "Ca_Cmol/Kg": "Calcio (Ca) (*)cmol(+)/Kg",
    "Mg_Cmol/Kg": "Magnesio (Mg) (*) cmol(+)/Kg",
    "Na_Cmol/Kg": "Sodio (Na) (*) cmol(+)/Kg",
    "K_Cmol/Kg": "Potasio (K) (*) cmol(+)/Kg",

    "Calcio Intercambiable": "Calcio (Ca) (*)cmol(+)/Kg",
    "Magnesio Intercambiable": "Magnesio (Mg) (*) cmol(+)/Kg",
    "Sodio Intercambiable": "Sodio (Na) (*) cmol(+)/Kg",
    "Potasio Intercambiable": "Potasio (K) (*) cmol(+)/Kg",

    "Carbonato de Calcio Equivalente": "CaCO3 _% Equivalente",
    "Materia Orgánica (AS-07 Método de Walkley y Black)": "MO_%",
    "Materia Organica (AS-07 Walkley y Black)": "MO_%",
    "Materia Orgánica por LECO": "MO_%",

    "Conductividad Electrica (Suelo)": "CE_mS/m",
    "pH. (Suelo)": "pH",

    "Fósforo Disponible (Bray y Kurtz)": "P_mg/kg)",
    "Fósforo Disponible Bray mpaes": "P_mg/kg)",
    "Potasio Disponible (MPAES)": "K_ppm",
    "Potasio Disponible (AA)": "K_ppm",

    "Arena": "Arena",
    "Arcilla": "Arcilla",
    "Limo": "Limo",
    "Clase Textural": "Clase Textural",

    "Nitrógeno Total": "N_%",
    "Nitrógeno Total Kjeldahl": "N_%",
}


# ------------------------------------------------------------
# 4. Write mapped values from database
# ------------------------------------------------------------

for target_header, input_header in mapping.items():
    value = get(person_row, input_header)

    if value is None:
        value = 0
        print(f"{target_header}: DB VALUE NOT FOUND or EMPTY -> set to 0")

    set_target_value(
        target_ws,
        target_headers,
        target_header,
        value,
        target_row_number,
    )

    print(f"{target_header}: {value}")


# ------------------------------------------------------------
# 5. Fixed values and user-selected crop
# ------------------------------------------------------------

fixed_values = {
    "Número de Muestras": 1,
    "Presentación Muestra": "1 kg",
    "Plan de Recomendación de Fertilización": USER_PRODUCT,
}

for target_header, value in fixed_values.items():
    ok = set_target_value(
        target_ws,
        target_headers,
        target_header,
        value,
        target_row_number,
    )

    if ok:
        print(f"{target_header}: {value}")


# ------------------------------------------------------------
# 6. Force specific target cells
# ------------------------------------------------------------

target_ws["T3"] = USER_PRODUCT
print(f"T3 / Plan de Recomendación de Fertilización: {USER_PRODUCT}")

target_ws["AL3"] = 0
target_ws["AM3"] = 0

if p_number is None:
    print("P_mg/kg) not found or invalid. AL3 and AM3 set to 0.")

elif ph_number is None:
    print("pH not found or invalid. Cannot choose Olsen/Bray. AL3 and AM3 set to 0.")

elif ph_number < 7.0:
    target_ws["AM3"] = p_number
    print(f"pH: {ph_number}")
    print("Using Bray because pH < 7.0")
    print("AL3 / Fósforo Disponible (Olsen): 0")
    print(f"AM3 / Fósforo Disponible (Bray y Kurtz): {p_number}")

else:
    target_ws["AL3"] = p_number
    print(f"pH: {ph_number}")
    print("Using Olsen because pH >= 7.0")
    print(f"AL3 / Fósforo Disponible (Olsen): {p_number}")
    print("AM3 / Fósforo Disponible (Bray y Kurtz): 0")


# ------------------------------------------------------------
# 7. Reinsert images
# ------------------------------------------------------------

reinsert_images(target_wb)


# ------------------------------------------------------------
# 8. Save output
# ------------------------------------------------------------

target_wb.save(OUTPUT_FILE)

print("=" * 70)
print("Done.")
print(f"Saved as: {OUTPUT_FILE.name}")
